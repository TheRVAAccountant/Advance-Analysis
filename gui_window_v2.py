import os
import logging
import shutil
import tkinter as tk
import threading
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook

class TextHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.see(tk.END)
        self.text_widget.update_idletasks()

class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Advance Analysis')
        self.geometry('600x500')
        self.configure(bg='#313131')
        self.resizable(True, True)

        self.grid_columnconfigure(1, weight=1)
        for i in range(7):
            self.grid_rowconfigure(i, weight=0)
        self.grid_rowconfigure(4, weight=1)

        self.apply_forest_dark_theme()
        self.initUI()

        icon_path = r"C:\Users\Jeron.Crooks\OneDrive - Department of Homeland Security\Documents\Python Scripts\Advance Analysis\money_suitcase.ico"
        self.iconbitmap(icon_path)

        self.logger = self.setup_logging()

    def apply_forest_dark_theme(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        tcl_file_path = os.path.join(script_dir, 'forest-dark.tcl')
        self.tk.call('source', tcl_file_path)
        ttk.Style().theme_use('forest-dark')

    def initUI(self):
        ttk.Label(self, text="Select Component Name:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.component_name_combo = ttk.Combobox(self, values=["CBP", "CG", "CIS", "CYB", "FEM", "FLE", "ICE", "MGA", "MGT", "OIG", "TSA", "SS", "ST", "WMD"], state="readonly")
        self.component_name_combo.set("WMD")
        self.component_name_combo.grid(row=0, column=1, columnspan=2, sticky="ew", padx=5, pady=5)

        ttk.Label(self, text="Advance Analysis File:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.target_file_edit = ttk.Entry(self)
        self.target_file_edit.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.target_file_edit)).grid(row=1, column=2, padx=5, pady=5)

        ttk.Label(self, text="Current Year Trial Balance File:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.cy_trial_balance_edit = ttk.Entry(self)
        self.cy_trial_balance_edit.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.cy_trial_balance_edit)).grid(row=2, column=2, padx=5, pady=5)

        ttk.Label(self, text="Prior Year Trial Balance File:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.py_trial_balance_edit = ttk.Entry(self)
        self.py_trial_balance_edit.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.py_trial_balance_edit)).grid(row=3, column=2, padx=5, pady=5)

        self.log_text = tk.Text(self, wrap=tk.WORD, width=70, height=10, bg='#232323', fg='white')
        self.log_text.grid(row=4, column=0, columnspan=3, padx=5, pady=5, sticky="nsew")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=4, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.progress_bar = ttk.Progressbar(self, orient="horizontal", length=580, mode="determinate")
        self.progress_bar.grid(row=5, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        ttk.Button(self, text="Start", command=self.start_operations).grid(row=6, column=0, columnspan=3, padx=5, pady=5)

    def browse_file(self, entry_widget):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)

    def start_operations(self):
        component_name = self.component_name_combo.get()
        target_file = self.target_file_edit.get()
        cy_trial_balance_file = self.cy_trial_balance_edit.get()
        py_trial_balance_file = self.py_trial_balance_edit.get()

        if not all([component_name, target_file, cy_trial_balance_file, py_trial_balance_file]):
            messagebox.showerror("Error", "Please select all required files and component name.")
            return

        self.progress_bar['value'] = 0
        self.update_idletasks()

        thread = OperationThread(
            component_name,
            target_file,
            cy_trial_balance_file,
            py_trial_balance_file,
            self.logger,
            self.update_progress
        )
        thread.start()

        self.monitor_thread(thread)

    def monitor_thread(self, thread):
        if thread.is_alive():
            self.after(100, lambda: self.monitor_thread(thread))
        else:
            if thread.exception:
                self.logger.error(f"An error occurred: {str(thread.exception)}")
                messagebox.showerror("Error", f"An error occurred: {str(thread.exception)}")
            else:
                self.logger.info("Operations completed successfully.")
                messagebox.showinfo("Complete", "Operations completed successfully!")

    def setup_logging(self):
        logger = logging.getLogger("MainLogger")
        logger.setLevel(logging.DEBUG)

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

        log_dir = "logs"
        os.makedirs(log_dir, exist_ok=True)
        log_filename = os.path.join(log_dir, f"AdvanceAnalysis_Log_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.txt")
        file_handler = logging.FileHandler(log_filename)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        text_handler = TextHandler(self.log_text)
        text_handler.setLevel(logging.DEBUG)
        text_handler.setFormatter(formatter)
        logger.addHandler(text_handler)

        return logger

    def update_progress(self, value):
        self.progress_bar['value'] = value
        self.update_idletasks()

class OperationThread(threading.Thread):
    def __init__(self, component_name, target_file, cy_trial_balance_file, py_trial_balance_file, logger, progress_callback):
        super().__init__()
        self.component_name = component_name
        self.target_file = target_file
        self.cy_trial_balance_file = cy_trial_balance_file
        self.py_trial_balance_file = py_trial_balance_file
        self.logger = logger
        self.progress_callback = progress_callback
        self.exception = None

    def run(self):
        try:
            self.logger.info("Operation started...")

            new_target_file = self.create_copy_of_target_file()

            if not self.copy_and_rename_sheet(self.cy_trial_balance_file, f"{self.component_name} Total", new_target_file, "DO CY TB", insert_index=3):
                self.logger.error(f"Failed to copy sheet CY '{self.component_name} Total'.")
                return

            if not self.copy_and_rename_sheet(self.py_trial_balance_file, f"{self.component_name} Total", new_target_file, "DO PY TB", insert_index=4):
                self.logger.error(f"Failed to copy sheet PY '{self.component_name} Total'.")
                return

            self.progress_callback(100)
            self.logger.info("Operation completed successfully.")
        except Exception as e:
            self.logger.error(f"Error during operation: {e}", exc_info=True)
            self.exception = e

    def create_copy_of_target_file(self):
        try:
            file_name, file_extension = os.path.splitext(self.target_file)
            new_file_name = f"{file_name} - DO{file_extension}"
            shutil.copy2(self.target_file, new_file_name)
            self.logger.info(f"Created copy of target file: {new_file_name}")
            return new_file_name
        except Exception as e:
            self.logger.error(f"Failed to create copy of target file: {e}")
            raise

    def copy_and_rename_sheet(self, source_path, source_sheet_name, target_path, new_sheet_name, insert_index=None):
        try:
            self.logger.info(f"Loading source workbook: {source_path}")
            source_wb = load_workbook(source_path, data_only=False)
            if source_sheet_name not in source_wb.sheetnames:
                self.logger.error(f"Sheet '{source_sheet_name}' not found in {source_path}")
                return False

            self.logger.info(f"Loading target workbook: {target_path}")
            target_wb = load_workbook(target_path, data_only=False)
            self.logger.info(f"Copying sheet '{source_sheet_name}' from source to target")
            source_sheet = source_wb[source_sheet_name]
            
            if insert_index is not None:
                target_sheet = target_wb.create_sheet(new_sheet_name, insert_index)
            else:
                target_sheet = target_wb.create_sheet(new_sheet_name)

            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.data_type == "f":
                        target_cell.value = cell.value
                    if cell.has_style:
                        self.copy_cell_style(cell, target_cell)

            self.copy_dimensions(source_sheet, target_sheet)

            self.logger.info(f"Saving changes to target workbook: {target_path}")
            target_wb.save(target_path)
            target_wb.close()

            self.logger.info(f"Successfully copied and renamed sheet to '{new_sheet_name}' with formatting and formulas preserved")
            return True

        except Exception as e:
            self.logger.error(f"An error occurred while copying sheet: {e}", exc_info=True)
            return False

    def copy_cell_style(self, source_cell, target_cell):
        try:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            target_cell.alignment = source_cell.alignment.copy()
        except Exception as e:
            raise RuntimeError(f"Failed to copy cell style: {e}")

    def copy_dimensions(self, source_sheet, target_sheet):
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = value.width
            target_sheet.column_dimensions[key].hidden = value.hidden

        for key, value in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[key].height = value.height
            target_sheet.row_dimensions[key].hidden = value.hidden

def main():
    app = MainWindow()
    app.mainloop()

if __name__ == "__main__":
    main()