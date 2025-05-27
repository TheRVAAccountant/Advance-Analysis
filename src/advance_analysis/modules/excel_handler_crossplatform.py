"""
Cross-platform Excel handling functionality for advance analysis.

This module provides functions for Excel file operations that work on all platforms,
including copying sheets with formatting and formulas preserved.
"""
import os
import logging
from typing import Optional, List, Tuple
import shutil
import re

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle, Protection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from ..utils.logging_config import get_logger

logger = get_logger(__name__)


def copy_excel_with_formatting(source_path: str, dest_path: str) -> None:
    """
    Copy an Excel file preserving all formatting, formulas, and structure.
    
    Args:
        source_path: Path to the source Excel file
        dest_path: Path to the destination Excel file
    """
    try:
        # Use shutil.copy2 to preserve metadata
        shutil.copy2(source_path, dest_path)
        logger.info(f"Excel file copied with formatting from {source_path} to {dest_path}")
    except Exception as e:
        logger.error(f"Error copying Excel file: {str(e)}", exc_info=True)
        raise


def find_sheet_with_component_total(workbook: Workbook, component: str) -> Optional[str]:
    """
    Find the sheet name that contains the component name and 'Total'.
    
    Args:
        workbook: The openpyxl workbook object
        component: The component name (e.g., "WMD", "CBP")
    
    Returns:
        The name of the matching sheet or None if not found
    """
    try:
        pattern = f"{component}.*Total"
        for sheet_name in workbook.sheetnames:
            if re.search(pattern, sheet_name, re.IGNORECASE):
                logger.info(f"Found sheet matching pattern '{pattern}': {sheet_name}")
                return sheet_name
        
        logger.warning(f"No sheet found matching pattern '{pattern}'")
        return None
    except Exception as e:
        logger.error(f"Error finding sheet with component total: {str(e)}", exc_info=True)
        return None


def copy_sheet_to_workbook(
    source_wb: Workbook, 
    dest_wb: Workbook, 
    source_sheet_name: str, 
    new_sheet_name: str,
    insert_after: Optional[str] = None
) -> None:
    """
    Copy a sheet from one workbook to another, preserving formatting and formulas.
    
    Args:
        source_wb: Source workbook
        dest_wb: Destination workbook
        source_sheet_name: Name of the sheet to copy
        new_sheet_name: New name for the copied sheet
        insert_after: Name of the sheet after which to insert the new sheet
    """
    try:
        if source_sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Source sheet '{source_sheet_name}' not found")
        
        source_sheet = source_wb[source_sheet_name]
        
        # Create new sheet in destination workbook
        if new_sheet_name in dest_wb.sheetnames:
            logger.warning(f"Sheet '{new_sheet_name}' already exists, it will be replaced")
            del dest_wb[new_sheet_name]
        
        # Create the new sheet
        dest_sheet = dest_wb.create_sheet(title=new_sheet_name)
        
        # Copy all cell data, formatting, and formulas
        copy_sheet_data(source_sheet, dest_sheet)
        
        # Move sheet to the correct position if needed
        if insert_after and insert_after in dest_wb.sheetnames:
            sheets = dest_wb.sheetnames
            insert_index = sheets.index(insert_after) + 1
            dest_wb.move_sheet(new_sheet_name, offset=insert_index - sheets.index(new_sheet_name))
        
        logger.info(f"Sheet '{source_sheet_name}' copied to '{new_sheet_name}' successfully")
        
    except Exception as e:
        logger.error(f"Error copying sheet: {str(e)}", exc_info=True)
        raise


def copy_sheet_data(source_sheet: Worksheet, dest_sheet: Worksheet) -> None:
    """
    Copy all data, formatting, and formulas from source to destination sheet.
    
    Args:
        source_sheet: Source worksheet
        dest_sheet: Destination worksheet
    """
    try:
        # Copy cell dimensions
        for row_num, row_dim in source_sheet.row_dimensions.items():
            dest_sheet.row_dimensions[row_num].height = row_dim.height
            dest_sheet.row_dimensions[row_num].hidden = row_dim.hidden
        
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            dest_sheet.column_dimensions[col_letter].width = col_dim.width
            dest_sheet.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Copy all cells with data, formulas, and formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                dest_cell = dest_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy value or formula
                if cell.data_type == 'f':
                    dest_cell.value = cell.value
                    dest_cell.data_type = 'f'
                else:
                    dest_cell.value = cell.value
                
                # Copy formatting
                if cell.has_style:
                    dest_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                        underline=cell.font.underline,
                        strike=cell.font.strike
                    )
                    dest_cell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                    dest_cell.border = Border(
                        left=Side(style=cell.border.left.style, color=cell.border.left.color),
                        right=Side(style=cell.border.right.style, color=cell.border.right.color),
                        top=Side(style=cell.border.top.style, color=cell.border.top.color),
                        bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color)
                    )
                    dest_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                        text_rotation=cell.alignment.text_rotation,
                        indent=cell.alignment.indent
                    )
                    dest_cell.number_format = cell.number_format
                    dest_cell.protection = Protection(
                        locked=cell.protection.locked,
                        hidden=cell.protection.hidden
                    )
        
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            dest_sheet.merge_cells(str(merged_range))
        
        # Copy page setup
        dest_sheet.page_setup.orientation = source_sheet.page_setup.orientation
        dest_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
        dest_sheet.page_setup.fitToPage = source_sheet.page_setup.fitToPage
        dest_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
        dest_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
        
        # Copy print options
        dest_sheet.print_options.horizontalCentered = source_sheet.print_options.horizontalCentered
        dest_sheet.print_options.verticalCentered = source_sheet.print_options.verticalCentered
        
        # Copy freeze panes
        dest_sheet.freeze_panes = source_sheet.freeze_panes
        
    except Exception as e:
        logger.error(f"Error copying sheet data: {str(e)}", exc_info=True)
        raise


def process_excel_files_crossplatform(
    output_path: str,
    input_path: str,
    current_dhstier_path: str,
    prior_dhstier_path: str,
    component: str,
    password: Optional[str] = None
) -> None:
    """
    Process Excel files by copying the Advance Analysis file and adding TB sheets.
    
    Args:
        output_path: Path to the processed output file (Review file)
        input_path: Path to the renamed Advance Analysis file  
        current_dhstier_path: Path to the current year DHSTIER Trial Balance file
        prior_dhstier_path: Path to the prior year DHSTIER Trial Balance file
        component: The component name (e.g., "WMD", "CBP")
        password: Password for protected sheets (not used in cross-platform version)
    """
    try:
        logger.info(f"Starting cross-platform Excel file processing for {component}")
        
        # First, copy the original Advance Analysis file to preserve all formatting and formulas
        # The input_path is already the renamed file, so we just need to work with it
        
        # Load workbooks
        logger.info("Loading workbooks...")
        dest_wb = openpyxl.load_workbook(input_path, data_only=False, keep_vba=True)
        current_dhstier_wb = openpyxl.load_workbook(current_dhstier_path, data_only=False, keep_vba=True)
        prior_dhstier_wb = openpyxl.load_workbook(prior_dhstier_path, data_only=False, keep_vba=True)
        
        # Find the sheet to insert after (typically the last sheet)
        insert_after_sheet = dest_wb.sheetnames[-1] if dest_wb.sheetnames else None
        
        # Find and copy current year DHSTIER sheet
        logger.info("Finding and copying current year DHSTIER sheet...")
        current_sheet_name = find_sheet_with_component_total(current_dhstier_wb, component)
        if current_sheet_name:
            copy_sheet_to_workbook(
                current_dhstier_wb,
                dest_wb,
                current_sheet_name,
                "CY DO TB",
                insert_after=insert_after_sheet
            )
            insert_after_sheet = "CY DO TB"
        else:
            logger.warning(f"Could not find current year sheet for component {component}")
        
        # Find and copy prior year DHSTIER sheet
        logger.info("Finding and copying prior year DHSTIER sheet...")
        prior_sheet_name = find_sheet_with_component_total(prior_dhstier_wb, component)
        if prior_sheet_name:
            copy_sheet_to_workbook(
                prior_dhstier_wb,
                dest_wb,
                prior_sheet_name,
                "PY DO TB",
                insert_after=insert_after_sheet
            )
        else:
            logger.warning(f"Could not find prior year sheet for component {component}")
        
        # Save the workbook
        logger.info(f"Saving workbook to {input_path}")
        dest_wb.save(input_path)
        dest_wb.close()
        
        # Close other workbooks
        current_dhstier_wb.close()
        prior_dhstier_wb.close()
        
        logger.info(f"Cross-platform Excel processing completed successfully for {component}")
        
    except Exception as e:
        logger.error(f"Error in cross-platform Excel processing: {str(e)}", exc_info=True)
        raise