"""
Excel handling functionality for obligation analysis.

This module provides functions for interacting with Excel files,
including formatting, copying sheets, and manipulating data.
"""
import os
import logging
import time
from typing import Optional, Any, Dict, List, Tuple
import re

import pandas as pd

# Try to import Windows-specific modules
try:
    import pythoncom
    import win32com.client
    from win32com.client import constants
    WINDOWS_COM_AVAILABLE = True
except ImportError:
    WINDOWS_COM_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("Windows COM modules not available - Excel COM automation features will be disabled")

# Import the new Excel processor if available
try:
    from .excel_processor import ExcelProcessor, safe_excel_operation as safe_excel_op
    EXCEL_PROCESSOR_AVAILABLE = True
except ImportError:
    EXCEL_PROCESSOR_AVAILABLE = False
    safe_excel_op = lambda f: f  # No-op decorator if processor not available

from ..utils.logging_config import get_logger
# from ..utils.helpers import format_currency, format_excel_style
from ..modules.file_handler import ensure_file_accessibility
# from ..core.udo_validation import validate_udo_tier_recon

# Enhanced file and COM handling utilities
def wait_for_file_ready(file_path: str, max_wait: float = 10.0, check_interval: float = 0.5) -> bool:
    """
    Wait for file to be accessible and not locked.
    
    Args:
        file_path: Path to the file to check
        max_wait: Maximum time to wait in seconds
        check_interval: Time between checks in seconds
        
    Returns:
        True if file is ready, False if timeout reached
    """
    import time
    start_time = time.time()
    
    while time.time() - start_time < max_wait:
        try:
            # Try to open file exclusively to check if it's locked
            with open(file_path, 'rb+') as f:
                f.read(1)
            logger.debug(f"File {file_path} is ready for access")
            return True
        except (IOError, OSError) as e:
            logger.debug(f"File {file_path} not ready: {e}")
            time.sleep(check_interval)
    
    logger.warning(f"Timeout waiting for file {file_path} to be ready")
    return False


def initialize_excel_com(max_retries: int = 3) -> Any:
    """
    Initialize Excel COM with validation and retry logic.
    
    Args:
        max_retries: Maximum number of initialization attempts
        
    Returns:
        Excel COM object
        
    Raises:
        Exception: If initialization fails after all retries
    """
    import time
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Initializing Excel COM (attempt {attempt + 1}/{max_retries})")
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            
            # Validate COM object
            if excel and hasattr(excel, 'Version'):
                excel.Visible = False
                excel.DisplayAlerts = False
                logger.info(f"Excel COM initialized successfully (Version: {excel.Version})")
                return excel
            else:
                raise Exception("Invalid Excel COM object")
                
        except Exception as e:
            logger.error(f"Failed to initialize Excel COM (attempt {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
                time.sleep(1)
            else:
                raise Exception(f"Failed to initialize Excel COM after {max_retries} attempts: {e}")


def open_workbook_robust(excel: Any, file_path: str, max_retries: int = 3, read_only: bool = False) -> Any:
    """
    Open workbook with retry logic and diagnostics.
    
    Args:
        excel: Excel COM application object
        file_path: Path to the workbook
        max_retries: Maximum number of open attempts
        read_only: Whether to open in read-only mode
        
    Returns:
        Workbook COM object
        
    Raises:
        Exception: If workbook cannot be opened after all retries
    """
    import time
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Opening workbook {file_path} (attempt {attempt + 1}/{max_retries})")
            
            # Ensure file is ready
            if not wait_for_file_ready(file_path):
                raise Exception(f"File not ready after timeout: {file_path}")
            
            # Clear any existing references
            excel.CutCopyMode = False
            
            # Open workbook with specific parameters to avoid issues
            wb = excel.Workbooks.Open(
                file_path,
                UpdateLinks=0,  # Don't update external links
                ReadOnly=read_only,
                IgnoreReadOnlyRecommended=True,
                Notify=False,
                AddToMru=False  # Don't add to recent files
            )
            
            # Validate workbook
            if wb and hasattr(wb, 'Sheets') and wb.Sheets.Count > 0:
                logger.info(f"Workbook opened successfully: {file_path} (Sheets: {wb.Sheets.Count})")
                return wb
            else:
                raise Exception("Invalid workbook object")
                
        except Exception as e:
            logger.warning(f"Failed to open workbook (attempt {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                time.sleep(2 * (attempt + 1))  # Exponential backoff
            else:
                raise Exception(f"Failed to open workbook after {max_retries} attempts: {file_path} - {e}")


def open_workbook_robust_v2(excel: Any, file_path: str, max_retries: int = 3, read_only: bool = False) -> Any:
    """
    Enhanced workbook opening with multiple strategies.
    
    Args:
        excel: Excel COM application object
        file_path: Path to the workbook
        max_retries: Maximum number of open attempts
        read_only: Whether to open in read-only mode
        
    Returns:
        Workbook COM object
        
    Raises:
        Exception: If workbook cannot be opened after all retries
    """
    import time
    
    # First ensure file is completely ready
    if not ensure_file_ready_after_copy(file_path, file_path):
        logger.warning(f"File not stabilized, attempting anyway: {file_path}")
    
    for attempt in range(max_retries):
        wb = None
        try:
            logger.info(f"Opening workbook {file_path} (attempt {attempt + 1}/{max_retries})")
            
            # Clear any existing references
            excel.CutCopyMode = False
            
            # Strategy 1: Open with minimal parameters
            try:
                logger.debug("Strategy 1: Minimal parameters")
                wb = excel.Workbooks.Open(file_path)
                if validate_workbook_robust(wb, file_path):
                    logger.info(f"Workbook opened successfully with strategy 1: {file_path}")
                    return wb
                else:
                    logger.debug("Strategy 1 validation failed")
                    if wb:
                        wb.Close(SaveChanges=False)
                        wb = None
            except Exception as e:
                logger.debug(f"Strategy 1 failed: {e}")
                if wb:
                    try:
                        wb.Close(SaveChanges=False)
                    except:
                        pass
                    wb = None
            
            # Strategy 2: Open with full parameters
            try:
                logger.debug("Strategy 2: Full parameters")
                wb = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=0,
                    ReadOnly=read_only,
                    IgnoreReadOnlyRecommended=True,
                    Notify=False,
                    AddToMru=False,
                    CorruptLoad=2  # xlRepairFile
                )
                if validate_workbook_robust(wb, file_path):
                    logger.info(f"Workbook opened successfully with strategy 2: {file_path}")
                    return wb
                else:
                    logger.debug("Strategy 2 validation failed")
                    if wb:
                        wb.Close(SaveChanges=False)
                        wb = None
            except Exception as e:
                logger.debug(f"Strategy 2 failed: {e}")
                if wb:
                    try:
                        wb.Close(SaveChanges=False)
                    except:
                        pass
                    wb = None
            
            # Strategy 3: Force Excel to refresh and retry
            if attempt < max_retries - 1:
                logger.debug("Strategy 3: Excel refresh")
                excel.ScreenUpdating = True
                excel.ScreenUpdating = False
                # Force Excel to process pending events
                try:
                    excel.Calculate()
                except:
                    pass
                time.sleep(2 * (attempt + 1))
                
        except Exception as e:
            logger.warning(f"All strategies failed (attempt {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                raise Exception(f"Failed to open workbook after {max_retries} attempts with all strategies: {file_path}")


def release_file_locks(file_path: str) -> None:
    """
    Ensure file is released from any locks before COM operations.
    
    Args:
        file_path: Path to the file to release
    """
    import gc
    import time
    
    logger.debug(f"Releasing file locks for: {file_path}")
    
    # Force garbage collection
    gc.collect()
    
    # Small delay to ensure file system catches up
    time.sleep(0.5)
    
    # Verify file exists and is accessible
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Try to ensure file is not locked
    try:
        with open(file_path, 'rb') as f:
            f.read(1)
        logger.debug(f"File {file_path} is accessible")
    except Exception as e:
        logger.warning(f"File may be locked: {file_path} - {e}")


def cleanup_com_objects(excel: Any, workbooks: List[Any]) -> None:
    """
    Enhanced cleanup with proper error handling.
    
    Args:
        excel: Excel COM application object
        workbooks: List of workbook COM objects to close
    """
    import time
    import gc
    
    logger.info("Starting enhanced COM cleanup process")
    
    # Clear clipboard
    if excel:
        try:
            excel.CutCopyMode = False
            logger.debug("Cleared Excel clipboard")
        except Exception as e:
            logger.debug(f"Could not clear clipboard: {e}")
    
    # Close workbooks
    for i, wb in enumerate(workbooks):
        if wb:
            try:
                wb.Close(SaveChanges=False)
                logger.debug(f"Closed workbook {i + 1}/{len(workbooks)}")
            except Exception as e:
                logger.debug(f"Error closing workbook {i + 1}: {e}")
            
            try:
                release_com_object(wb)
            except:
                pass
    
    # Quit Excel
    if excel:
        try:
            excel.Quit()
            logger.debug("Excel application quit successfully")
        except Exception as e:
            logger.debug(f"Error quitting Excel: {e}")
        
        try:
            release_com_object(excel)
        except:
            pass
    
    # Force cleanup
    gc.collect()
    time.sleep(0.5)
    
    # Uninitialize COM
    try:
        pythoncom.CoUninitialize()
        logger.debug("COM uninitialized")
    except Exception as e:
        logger.debug(f"Error uninitializing COM: {e}")
    
    logger.info("Enhanced COM cleanup completed")

# Temporary placeholder functions
def format_currency(value):
    """Format a value as currency."""
    try:
        return f"${float(value):,.2f}"
    except:
        return str(value)

def format_excel_style(worksheet, style_name):
    """Apply Excel style to worksheet (placeholder)."""
    pass

logger = get_logger(__name__)


class ExcelComContext:
    """Context manager for Excel COM operations to ensure proper cleanup."""
    
    def __init__(self):
        self.excel = None
        self.workbooks = []
        
    def __enter__(self):
        """Initialize COM and create Excel application."""
        try:
            pythoncom.CoInitialize()
            self.excel = win32com.client.Dispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            logger.debug("Excel COM context initialized")
            return self
        except Exception as e:
            logger.error(f"Failed to initialize Excel COM: {e}")
            raise
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Clean up Excel COM objects."""
        logger.debug("Cleaning up Excel COM context...")
        
        # Clear clipboard
        if self.excel:
            try:
                self.excel.CutCopyMode = False
            except:
                pass
        
        # Close all tracked workbooks
        for wb in self.workbooks:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
            release_com_object(wb)
        
        # Quit Excel
        if self.excel:
            try:
                self.excel.Quit()
            except:
                pass
            release_com_object(self.excel)
        
        # Force garbage collection
        import gc
        gc.collect()
        time.sleep(0.2)
        
        # Uninitialize COM
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        
        logger.debug("Excel COM context cleaned up")
        return False  # Don't suppress exceptions
    
    def open_workbook(self, path, **kwargs):
        """Open a workbook and track it for cleanup."""
        wb = self.excel.Workbooks.Open(path, **kwargs)
        self.workbooks.append(wb)
        return wb


# Error handling wrapper function
def safe_excel_operation(func):
    """
    Decorator for safely executing Excel operations and handling exceptions.
    
    Args:
        func: The function to decorate.
        
    Returns:
        The decorated function.
    """
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"Error in {func.__name__}: {str(e)}", exc_info=True)
            raise
    return wrapper


@safe_excel_operation
def format_excel_file(file_path: str) -> None:
    """
    Format the Excel file with specific styles and column adjustments.
    
    Args:
        file_path (str): Path to the Excel file to be formatted.
    
    Raises:
        FileNotFoundError: If the file doesn't exist.
        Exception: For other errors during formatting.
    """
    try:
        logger.info(f"Starting to format Excel file: {file_path}")
        
        # Read the Excel file
        df = pd.read_excel(file_path, engine='openpyxl')
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write the dataframe to the Excel file
            df.to_excel(writer, index=False, sheet_name='DO Tab 4 Review')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['DO Tab 4 Review']
            
            # Apply formatting
            _apply_excel_formatting(worksheet)

        logger.info(f"Excel file formatted successfully: {file_path}")
    except FileNotFoundError as e:
        logger.error(f"File not found: {file_path}")
        raise
    except Exception as e:
        logger.error(f"Error formatting Excel file: {str(e)}", exc_info=True)
        raise


def _apply_excel_formatting(worksheet) -> None:
    """
    Apply formatting to the worksheet including styles, borders, and column widths.
    
    Args:
        worksheet: The openpyxl worksheet object.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle, Protection
    
    # Define styles
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = '#,##0.00_);[Red](#,##0.00)'

    date_style = NamedStyle(name="date_style")
    date_style.number_format = 'mm/dd/yyyy'

    age_style = NamedStyle(name="age_style")
    age_style.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

    wrap_style = NamedStyle(name="wrap_style")
    wrap_style.alignment = Alignment(vertical="center", wrap_text=True)

    # Apply styles to header row
    for cell in worksheet[1]:
        cell.style = header_style

    # Get column indices and apply specific formats
    column_indices = {cell.value: cell.column_letter for cell in worksheet[1]}

    currency_columns = ["PY Q4 Ending Balance UDO", "Current FY Quarter-End  balance UDO", "Current FY Quarter-End  balance UDO_comp"]
    date_columns = ['Date of Obligation', 'Date of the Last Invoice Received', 'For Status 3 and 4 -Date deobligation was initiated', 
                    'For Status 3 and 4 - Date debligation is planned', 'Period of Performance End Date', 'Date Component Last Contacted Vendor for Bill',
                    'Date Component Last Contacted Vendor for Bill_comp', 'For Status 3 and 4 -Date deobligation was initiated_comp', 
                    'For Status 3 and 4 - Date debligation is planned_comp', 'Date of Obligation_comp', 
                    'Reporting Date', 'FY Start Date', 'FY End Date']
    age_columns = ['Age of Obligation in Days2', 'UDO Age in Days', 'De-Ob Date Change in Days']
    wrap_columns = ['Active / Inactive Obligation (No Invoice in Over 1 Year)', 'Null or Blank Columns', 'Prior Status Agrees?', 'DO Comment']

    # Apply column-specific styles
    for col_name, col_letter in column_indices.items():
        if col_name in currency_columns:
            for cell in worksheet[col_letter][1:]:
                cell.style = currency_style
        elif col_name in date_columns:
            for cell in worksheet[col_letter][1:]:
                cell.style = date_style
        elif col_name in age_columns:
            for cell in worksheet[col_letter][1:]:
                cell.style = age_style
        elif col_name in wrap_columns:
            for cell in worksheet[col_letter][1:]:
                cell.style = wrap_style

    # Apply borders to all cells
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if not cell.border:
                cell.border = border

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)

    # Freeze the header row
    worksheet.freeze_panes = "A2"


@safe_excel_operation
def advanced_copy_sheet(source_wb, dest_wb, source_sheet_name: str, new_sheet_name: Optional[str] = None, insert_after: Optional[str] = None) -> None:
    """
    Copy a sheet from one workbook to another, preserving formatting.
    
    Args:
        source_wb: Source workbook (win32com.client.CDispatch object)
        dest_wb: Destination workbook (win32com.client.CDispatch object)
        source_sheet_name (str): Name of the sheet to copy
        new_sheet_name (str, optional): New name for the copied sheet. If None, original name is kept.
        insert_after (str, optional): Name of the sheet after which to insert the new sheet. If None, inserts at the end.
    
    Raises:
        ValueError: If source sheet or insert_after sheet not found.
    """
    try:
        # Ensure the source sheet exists
        source_sheet = None
        for sheet in source_wb.Worksheets:
            if sheet.Name == source_sheet_name:
                source_sheet = sheet
                break
        if not source_sheet:
            raise ValueError(f"Source sheet '{source_sheet_name}' not found in source workbook.")

        # Ensure the insert_after sheet exists in the destination workbook
        if insert_after:
            insert_after_sheet = None
            for sheet in dest_wb.Worksheets:
                if sheet.Name == insert_after:
                    insert_after_sheet = sheet
                    break
            if not insert_after_sheet:
                raise ValueError(f"Insert-after sheet '{insert_after}' not found in destination workbook.")
            insert_position = insert_after_sheet.Index + 1
        else:
            insert_position = dest_wb.Worksheets.Count + 1

        # Temporarily make the source workbook active
        source_wb.Activate()

        # Copy the sheet to the destination workbook
        source_sheet.Copy(After=dest_wb.Worksheets(insert_position - 1))

        # Get the newly copied sheet in the destination workbook
        new_sheet = dest_wb.Worksheets(insert_position)

        # Rename the sheet if a new name is provided
        if new_sheet_name:
            new_sheet.Name = new_sheet_name

        logger.info(f"Sheet '{source_sheet_name}' copied successfully" + 
                    (f" and renamed to '{new_sheet_name}'" if new_sheet_name else "") +
                    f" at position {insert_position}")
    except ValueError as e:
        logger.error(f"Sheet copying error: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Error copying sheet '{source_sheet_name}': {str(e)}", exc_info=True)
        raise


@safe_excel_operation
def find_sheet_name(wb, component: str) -> str:
    """
    Find the sheet name that starts with the component name and ends with 'Total'.
    
    Args:
        wb: Workbook (win32com.client.CDispatch object)
        component (str): The component name (e.g., "WMD", "CBP")
    
    Returns:
        str: The name of the matching sheet
    
    Raises:
        ValueError: If no matching sheet is found.
    """
    sheet_name_start = f"{component} "
    for sheet in wb.Sheets:
        if sheet.Name.startswith(sheet_name_start) and sheet.Name.endswith("Total"):
            return sheet.Name
    raise ValueError(f"No sheet found starting with '{sheet_name_start}' and ending with 'Total'")


@safe_excel_operation
def create_tickmark_legend_and_compare_values(wb, password: str) -> None:
    """
    Creates the tickmark legend, compares PY Q4 Ending Balance values, and adds validation marks.
    
    Args:
        wb: Workbook (win32com.client.CDispatch object)
        password (str): Password to unprotect sheets
    """
    try:
        # Find the Certification sheet
        cert_sheet = None
        for sheet in wb.Sheets:
            if "Certification" in sheet.Name or "2-Certification" in sheet.Name:
                cert_sheet = sheet
                break
        
        if not cert_sheet:
            logger.error("Certification sheet not found in the workbook")
            return

        logger.info(f"Creating tickmark legend in sheet: {cert_sheet.Name}")
        
        # Define the tickmark data
        tickmarks = [
            ("Tickmark", "Calibri", 11, True, ""),
            ("a", "Marlett", 11, False, "Agreed to TB"),
            ("a", "Wingdings", 10, False, "Foot"),
            ("b", "Wingdings", 10, False, "Crossfoot"),
            ("h", "Wingdings", 10, False, "Agreed to Deliverable Support"),
            ("8", "Wingdings 2", 10, False, "Agreed to Reconciliation"),
            ("Rx", "Calibri", 10, False, "Recalculated/Verified formula"),
            ("i", "Wingdings", 10, False, "Agreed to TIER report"),
            ("X", "Calibri", 10, True, "Does Not Agree"),
            ("m", "Wingdings", 10, False, "Agreed to Certification")
        ]
        
        # Apply the tickmarks
        for i, (symbol, font_name, font_size, is_bold, description) in enumerate(tickmarks):
            cell_g = cert_sheet.Cells(i+1, 7)  # Column G
            cell_g.Value = symbol
            cell_g.Font.Name = font_name
            cell_g.Font.Size = font_size
            cell_g.Font.Bold = is_bold
            cell_g.Font.Color = 0 
            
            if description:
                cell_h = cert_sheet.Cells(i+1, 8)  # Column H
                cell_h.Value = description
                cell_h.Font.Name = "Calibri"
                cell_h.Font.Size = 11
                cell_h.Font.Color = 0
        
        # Apply formatting to the range
        range_to_format = cert_sheet.Range("G2:H10")
        range_to_format.Borders.LineStyle = win32com.client.constants.xlContinuous
        range_to_format.Borders.Weight = win32com.client.constants.xlThick
        range_to_format.HorizontalAlignment = win32com.client.constants.xlCenter
        range_to_format.VerticalAlignment = win32com.client.constants.xlCenter

        # Auto-fit column H
        cert_sheet.Columns("H").AutoFit()
        
        logger.info("Tickmark legend created successfully")

        # Find "Advances" in Column B of the Certification sheet
        logger.info("Searching for 'Advances' in Column B of Certification sheet")
        advances_cell = None
        
        # Search column B (column index 2) for "Advances"
        for row in range(1, 100):  # Search first 100 rows
            cell_value = cert_sheet.Cells(row, 2).Value
            if cell_value and "Advances" in str(cell_value):
                advances_cell = cert_sheet.Cells(row, 2)
                logger.info(f"Found 'Advances' in cell B{row}: '{cell_value}'")
                break
        
        if not advances_cell:
            logger.error("'Advances' not found in Column B of Certification sheet")
            return

        # The value is in the row immediately below the "Advances" cell
        cert_value_cell = cert_sheet.Cells(advances_cell.Row + 1, advances_cell.Column)
        logger.info(f"Using value from cell {cert_value_cell.Address} (row below 'Advances')")
        cert_value = cert_value_cell.Value
        formatted_cert_value = format_currency(cert_value)
        
        logger.info(f"Certification value found in cell {cert_value_cell.Address}: {formatted_cert_value}")
        
        # Log surrounding cells for debugging
        logger.debug(f"Cell above (B{advances_cell.Row}): {advances_cell.Value}")
        logger.debug(f"Cell value (B{cert_value_cell.Row}): {cert_value}")
        if cert_value_cell.Row < 100:
            logger.debug(f"Cell below (B{cert_value_cell.Row + 1}): {cert_sheet.Cells(cert_value_cell.Row + 1, 2).Value}")

        # Find the PY Q4 Ending Balance sheet
        py_q4_sheet = None
        for sheet in wb.Sheets:
            if "PY Q4 Ending Balance" in sheet.Name or "3-PY Q4 Ending Balance" in sheet.Name:
                py_q4_sheet = sheet
                break

        if not py_q4_sheet:
            logger.error("PY Q4 Ending Balance sheet not found")
            return

        logger.info(f"PY Q4 Ending Balance sheet found: {py_q4_sheet.Name}")

        # Unprotect the sheet
        py_q4_sheet.Unprotect(Password=password)

        # Find the "TAS" cell in column A
        tas_cell = py_q4_sheet.Cells.Find("TAS", After=py_q4_sheet.Cells(1, 1), LookIn=win32com.client.constants.xlValues, LookAt=win32com.client.constants.xlWhole)
        if not tas_cell:
            logger.error("TAS cell not found in column A")
            return

        logger.info(f"TAS cell found at row {tas_cell.Row}, column {tas_cell.Column}")

        header_row = tas_cell.Row

        # The sum cell should be in column I, one row below the header
        sum_cell = py_q4_sheet.Cells(header_row + 1, 9)  # Column I is 9
        logger.info(f"Sum cell located at {sum_cell.Address}")

        py_q4_value = sum_cell.Value
        formatted_py_q4_value = format_currency(py_q4_value)
        logger.info(f"PY Q4 value found: {formatted_py_q4_value}")

        # Type conversion and comparison
        try:
            cert_value = float(cert_value) if cert_value is not None else None
            py_q4_value = float(py_q4_value) if py_q4_value is not None else None
            formatted_cert_value = format_currency(cert_value)
            formatted_py_q4_value = format_currency(py_q4_value)
        except (ValueError, TypeError) as e:
            logger.error(f"Value conversion error: {e}")
            return

        if cert_value is None or py_q4_value is None:
            logger.error(f"One or both values are None. Cert value: {formatted_cert_value}, PY Q4 value: {formatted_py_q4_value}")
            return

        logger.info(f"Comparing values: Cert value ({formatted_cert_value}) vs PY Q4 value ({formatted_py_q4_value})")

        if abs(cert_value - py_q4_value) < 0.01:  # Using a small threshold for float comparison
            logger.info("Values match within threshold. Adding checkmarks.")
            # Add 'h' in Wingdings on Certification sheet
            tick_cell = cert_sheet.Cells(cert_value_cell.Row, cert_value_cell.Column + 1)
            tick_cell.Font.Name = "Wingdings"
            tick_cell.Font.Size = 10
            tick_cell.Value = "h"

            # Add 'm' in Wingdings on PY Q4 Ending Balance sheet
            py_q4_tick_cell = py_q4_sheet.Cells(header_row + 1, 10)  # Column J is 10
            py_q4_tick_cell.Font.Name = "Wingdings"
            py_q4_tick_cell.Font.Size = 10
            py_q4_tick_cell.Font.Color = 0  # Black
            py_q4_tick_cell.Value = "m"
            
            logger.info("Tickmarks added successfully.")
        else:
            logger.info("Values do not match. Adding 'X' marks.")
            # Add 'X' in bold Calibri on Certification sheet
            cert_x_cell = cert_sheet.Cells(cert_value_cell.Row, cert_value_cell.Column + 1)
            cert_x_cell.Font.Name = "Calibri"
            cert_x_cell.Font.Size = 11
            cert_x_cell.Font.Bold = True
            cert_x_cell.Value = "X"

            # Add 'X' in bold Calibri on PY Q4 Ending Balance sheet
            py_q4_x_cell = sum_cell.Offset(0, 1)
            py_q4_x_cell.Font.Name = "Calibri"
            py_q4_x_cell.Font.Size = 11
            py_q4_x_cell.Font.Bold = True
            py_q4_x_cell.Value = "X"

            logger.warning(f"Values do not match. Cert: {formatted_cert_value}, PY Q4: {formatted_py_q4_value}. 'X' marks added.")

    except Exception as e:
        logger.error(f"Error in create_tickmark_legend_and_compare_values: {str(e)}", exc_info=True)


@safe_excel_operation
def create_pivot_table(wb, password: str) -> str:
    """
    Creates a pivot table in the PY Q4 Ending Balance sheet.
    
    Args:
        wb: Workbook (win32com.client.CDispatch object)
        password (str): Password to unprotect sheets
    
    Returns:
        str: The address of the sum cell
    """
    try:
        logger.info("Creating pivot table in sheet: 3-PY Q4 Ending Balance")
        
        # Find the correct sheet
        target_sheet = None
        for sheet in wb.Sheets:
            if "PY Q4 Ending Balance" in sheet.Name or "3-PY Q4 Ending Balance" in sheet.Name:
                target_sheet = sheet
                break
        
        if not target_sheet:
            raise ValueError("PY Q4 Ending Balance sheet not found")
        
        # Unprotect the sheet
        target_sheet.Unprotect(Password=password)
        
        # Find the "TAS" cell in column A
        tas_cell = target_sheet.Cells.Find("TAS", After=target_sheet.Cells(1, 1), LookIn=win32com.client.constants.xlValues, LookAt=win32com.client.constants.xlWhole)
        if not tas_cell:
            raise ValueError("TAS cell not found in column A")
        
        logger.info(f"TAS cell found at row {tas_cell.Row}, column {tas_cell.Column}")
        
        header_row = tas_cell.Row
        
        # Find the last column
        last_column = target_sheet.Cells(header_row, target_sheet.Columns.Count).End(win32com.client.constants.xlToLeft).Column
        logger.info(f"Last populated column in header row: {get_column_letter(last_column)}")
        
        # Log all header row values for debugging
        logger.info("Header row values in '3-PY Q4 Ending Balance' sheet:")
        header_values = []
        for col in range(1, last_column + 1):
            cell_value = target_sheet.Cells(header_row, col).Value
            col_letter = get_column_letter(col)
            header_values.append(f"{col_letter}: {cell_value}")
            logger.info(f"  Column {col_letter} (index {col}): {cell_value}")
        logger.info(f"Total columns with headers: {len(header_values)}")
        
        # Find the last row
        last_row = max(target_sheet.Cells(target_sheet.Rows.Count, col).End(win32com.client.constants.xlUp).Row 
                       for col in range(1, last_column + 1))
        logger.info(f"Last populated row after header row: {last_row}")
        
        # Define the data range
        data_range = target_sheet.Range(target_sheet.Cells(header_row, 1), target_sheet.Cells(last_row, last_column))
        logger.info(f"Pivot table range: {data_range.Address}")
        
        # Check if Advance/Prepayment column exists in the header row
        advance_col_index = None
        logger.info("Searching for Advance/Prepayment column...")
        for col in range(1, last_column + 1):
            cell_value = target_sheet.Cells(header_row, col).Value
            if cell_value:
                # Check for exact match and variations
                cell_str = str(cell_value).strip()
                # Check for Advance/Prepayment in various formats
                if "advance" in cell_str.lower() and "prepayment" in cell_str.lower():
                    advance_col_index = col
                    logger.info(f"Found Advance/Prepayment column at index {col} (column {get_column_letter(col)}): '{cell_value}'")
                    break
        
        if not advance_col_index:
            logger.warning("Advance/Prepayment column not found in header row. Looking for alternative balance columns...")
            # Find columns that might contain balance data
            balance_keywords = ['balance', 'amount', 'prepayment', 'advance']
            balance_col = None
            for col in range(1, last_column + 1):
                cell_value = str(target_sheet.Cells(header_row, col).Value or '').lower()
                if any(keyword in cell_value for keyword in balance_keywords):
                    balance_col = col
                    logger.info(f"Using column {get_column_letter(col)} ({target_sheet.Cells(header_row, col).Value}) for sum calculation")
                    break
            
            if balance_col:
                advance_col_index = balance_col
            else:
                logger.error("Could not find any Advance/Prepayment or balance column to sum")
                raise ValueError("No Advance/Prepayment or balance column found for calculation")
        
        # Get the column name for the pivot table
        advance_col_name = target_sheet.Cells(header_row, advance_col_index).Value
        logger.info(f"Using column '{advance_col_name}' for pivot table")
        
        # Try to create pivot table
        try:
            logger.info("Creating pivot table...")
            logger.info(f"Pivot table data range: {data_range.Address}")
            logger.info(f"Pivot table destination: Cell I{header_row} (column 9)")
            
            # Create pivot cache
            pivot_cache = wb.PivotCaches().Create(SourceType=win32com.client.constants.xlDatabase, SourceData=data_range)
            logger.info("Pivot cache created successfully")
            
            # Create pivot table
            pivot_table = pivot_cache.CreatePivotTable(TableDestination=target_sheet.Cells(header_row, 9), TableName="PYQ4BalancePivot")
            logger.info("Pivot table created successfully")
            
            # Add Advance/Prepayment field to values
            logger.info(f"Adding '{advance_col_name}' field to pivot table values...")
            
            # List all available pivot fields for debugging
            logger.info("Available pivot fields:")
            try:
                for i in range(1, pivot_table.PivotFields().Count + 1):
                    field_name = pivot_table.PivotFields(i).Name
                    logger.info(f"  Field {i}: {field_name}")
            except Exception as list_error:
                logger.warning(f"Could not list pivot fields: {str(list_error)}")
            
            # Try to get the field
            advance_field = pivot_table.PivotFields(advance_col_name)
            
            # Attempt to set the Function property, with fallback options
            try:
                advance_field.Orientation = win32com.client.constants.xlDataField
                advance_field.Function = win32com.client.constants.xlSum
                logger.info("Successfully set pivot field orientation and function")
            except Exception as e:
                logger.warning(f"Error setting field properties directly: {str(e)}")
                try:
                    # Alternative method: Add the field and then set properties
                    pivot_table.AddDataField(advance_field, f"Sum of {advance_col_name}", win32com.client.constants.xlSum)
                    logger.info("Successfully added data field using AddDataField method")
                except Exception as e2:
                    logger.error(f"Failed to add field using alternative method: {str(e2)}")
                    raise
            
            # Format the sum cell
            sum_cell = target_sheet.Cells(header_row + 1, 9)
            sum_cell.NumberFormat = "$#,##0.00_);[Red]($#,##0.00)"
            logger.info(f"Pivot table created successfully with sum in cell {sum_cell.Address}")
            
        except Exception as pivot_error:
            logger.warning(f"Failed to create pivot table: {str(pivot_error)}. Falling back to manual SUM formula.")
            # Fallback: Create a SUM formula for the Advance/Prepayment column
            sum_cell = target_sheet.Cells(header_row + 1, 9)
            col_letter = get_column_letter(advance_col_index)
            sum_cell.Formula = f"=SUM({col_letter}{header_row + 1}:{col_letter}{last_row})"
            sum_cell.NumberFormat = "$#,##0.00_);[Red]($#,##0.00)"
            target_sheet.Cells(header_row, 9).Value = "Total Advance/Prepayment"
            target_sheet.Cells(header_row, 9).Font.Bold = True
            logger.info(f"Created manual sum formula for Advance/Prepayment in cell {sum_cell.Address}")
        
        # Auto-fit column I
        target_sheet.Columns("I").AutoFit()
        
        # Return the sum_cell address
        if 'sum_cell' not in locals():
            sum_cell = target_sheet.Cells(header_row + 1, 9)
             
        logger.info("Pivot table or summary calculation created successfully")
        sum_cell_address = sum_cell.Address
        logger.info(f"Sum cell address: {sum_cell_address}")
        return sum_cell_address
        
    except Exception as e:
        logger.error(f"Error creating pivot table: {str(e)}", exc_info=True)
        raise


def apply_date_formatting(sheet: Any) -> None:
    """
    Apply date formatting to date columns in a sheet.
    
    Args:
        sheet: Worksheet COM object
    """
    date_columns = [
        'Date of Advance', 'Last Activity Date', 'Anticipated Liquidation Date',
        'Period of Performance End Date', 'Date of Advance_comp', 
        'Last Activity Date_comp', 'Anticipated Liquidation Date_comp',
        'Period of Performance End Date_comp'
    ]
    
    try:
        # Find header row (assume row 1)
        header_row = 1
        used_range = sheet.UsedRange
        
        if not used_range:
            return
            
        # Check each column header
        for col in range(1, used_range.Columns.Count + 1):
            try:
                header_value = sheet.Cells(header_row, col).Value
                if header_value in date_columns:
                    # Apply date format to entire column
                    col_letter = get_column_letter(col)
                    last_row = used_range.Rows.Count
                    date_range = sheet.Range(f"{col_letter}2:{col_letter}{last_row}")
                    date_range.NumberFormat = "m/d/yyyy"
                    logger.debug(f"Applied date formatting to column {col_letter}")
            except:
                pass
                
    except Exception as e:
        logger.warning(f"Error applying date formatting: {e}")


def populate_do_tab_4_review_sheet_pandas(input_wb: Any, dataframe_path: str) -> None:
    """
    Populate DO Tab 4 Review sheet using pandas as fallback.
    
    Args:
        input_wb: Input workbook COM object
        dataframe_path: Path to the dataframe Excel file
    """
    try:
        import pandas as pd
        
        # Read the dataframe
        df = pd.read_excel(dataframe_path, sheet_name=0, engine='openpyxl')
        
        # Get the DO Tab 4 Review sheet
        target_sheet = input_wb.Sheets("DO Tab 4 Review")
        
        # Clear existing data
        target_sheet.UsedRange.ClearContents()
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            target_sheet.Cells(1, col_idx).Value = str(col_name)
        
        # Write data
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                if pd.notna(value):
                    if isinstance(value, pd.Timestamp):
                        target_sheet.Cells(row_idx + 2, col_idx).Value = value.to_pydatetime()
                    else:
                        target_sheet.Cells(row_idx + 2, col_idx).Value = value
        
        # Apply date formatting
        apply_date_formatting(target_sheet)
        
        logger.info("DO Tab 4 Review sheet populated successfully using pandas")
        
    except Exception as e:
        logger.error(f"Error in pandas fallback: {e}")
        raise


def get_column_letter(column_number: int) -> str:
    """
    Convert a column number to a column letter (A, B, C, ..., Z, AA, AB, ...).
    
    Args:
        column_number (int): The column number (1-based).
    
    Returns:
        str: The column letter.
    """
    dividend = column_number
    column_letter = ''
    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        dividend = (dividend - modulo) // 26
    return column_letter


def release_com_object(obj):
    """
    Safely release a COM object and clean up references.
    
    Args:
        obj: COM object to release
    """
    if obj is not None:
        try:
            # Try to delete the object
            del obj
        except:
            pass
        
        # Force garbage collection
        import gc
        gc.collect()


def validate_com_object(obj, object_type: str = "generic") -> bool:
    """
    Validate that a COM object is properly initialized and connected.
    
    Args:
        obj: COM object to validate
        object_type: Type of object (e.g., "Excel", "Workbook", "Sheet")
        
    Returns:
        True if object is valid, False otherwise
    """
    if obj is None:
        logger.debug(f"COM object {object_type} is None")
        return False
    
    try:
        # Test basic COM functionality
        if object_type == "Excel":
            # Test Excel-specific property
            version = obj.Version
            logger.debug(f"Excel COM object validated (Version: {version})")
            return True
        elif object_type == "Workbook":
            # Test workbook-specific property
            sheet_count = obj.Sheets.Count
            logger.debug(f"Workbook COM object validated (Sheets: {sheet_count})")
            return True
        elif object_type == "Sheet":
            # Test sheet-specific property
            name = obj.Name
            logger.debug(f"Sheet COM object validated (Name: {name})")
            return True
        else:
            # Generic COM test
            str(obj)
            logger.debug(f"COM object {object_type} appears valid")
            return True
            
    except Exception as e:
        logger.debug(f"COM object {object_type} validation failed: {e}")
        return False


def ensure_com_connected(excel) -> bool:
    """
    Ensure Excel COM object is still connected and responsive.
    
    Args:
        excel: Excel COM application object
        
    Returns:
        True if connected, False otherwise
    """
    try:
        # Try a simple property access
        excel.DisplayAlerts = excel.DisplayAlerts
        return True
    except Exception as e:
        logger.warning(f"Excel COM object disconnected: {e}")
        return False


def validate_workbook_robust(wb: Any, file_path: str, max_wait: float = 5.0) -> bool:
    """
    Robustly validate workbook with retry for lazy initialization.
    
    Args:
        wb: Workbook COM object
        file_path: Path to the workbook file
        max_wait: Maximum time to wait for validation
        
    Returns:
        True if workbook is valid, False otherwise
    """
    import time
    start_time = time.time()
    
    while time.time() - start_time < max_wait:
        try:
            # Try multiple validation approaches
            if wb is None:
                return False
                
            # Try to access Name property first (lightweight)
            try:
                name = wb.Name
                logger.debug(f"Workbook name accessible: {name}")
            except:
                logger.debug("Workbook name not accessible")
                time.sleep(0.5)
                continue
            
            # Try to count sheets with error handling
            try:
                sheet_count = wb.Sheets.Count
                if sheet_count > 0:
                    logger.debug(f"Workbook validated: {sheet_count} sheets")
                    return True
            except Exception as e:
                logger.debug(f"Sheet count failed: {e}")
                
            # Alternative: Try to access first sheet
            try:
                first_sheet = wb.Sheets(1)
                if first_sheet:
                    logger.debug("Workbook validated via first sheet access")
                    return True
            except:
                pass
                
            time.sleep(0.5)
            
        except Exception as e:
            logger.debug(f"Validation attempt failed: {e}")
            time.sleep(0.5)
    
    return False


def ensure_file_ready_after_write(file_path: str, expected_size: Optional[int] = None, max_wait: float = 10.0) -> bool:
    """
    Ensure file is ready after write operation (copy or save).
    
    Args:
        file_path: Path to the file to check
        expected_size: Expected file size if known
        max_wait: Maximum time to wait
        
    Returns:
        True if file is ready, False otherwise
    """
    import time
    import os
    
    if not os.path.exists(file_path):
        return False
    
    # First, ensure basic file accessibility
    if not wait_for_file_ready(file_path, max_wait=max_wait/2):
        logger.warning(f"File not accessible: {file_path}")
        return False
    
    # If we know the expected size, wait for it
    if expected_size is not None:
        start_time = time.time()
        while time.time() - start_time < max_wait/2:
            try:
                current_size = os.path.getsize(file_path)
                if current_size == expected_size:
                    logger.debug(f"File size matches expected: {current_size} bytes")
                    return True
            except:
                pass
            time.sleep(0.5)
    
    # Otherwise, use the general ready check
    return ensure_file_ready_after_copy(file_path, file_path, max_wait=max_wait/2)


def ensure_file_ready_after_copy(source_path: str, dest_path: str, max_wait: float = 10.0) -> bool:
    """
    Ensure file is ready after copy operation with size stability check.
    
    Args:
        source_path: Source file path (for reference)
        dest_path: Destination file path to check
        max_wait: Maximum time to wait
        
    Returns:
        True if file is ready, False otherwise
    """
    import time
    import os
    
    if not os.path.exists(dest_path):
        return False
    
    # Wait for file size to stabilize
    last_size = -1
    stable_count = 0
    check_interval = 0.5
    
    start_time = time.time()
    while time.time() - start_time < max_wait:
        try:
            current_size = os.path.getsize(dest_path)
            
            if current_size == last_size:
                stable_count += 1
                if stable_count >= 3:  # Size stable for 3 checks
                    logger.debug(f"File size stable at {current_size} bytes")
                    # Additional check - can we open it?
                    with open(dest_path, 'rb') as f:
                        f.read(1)
                    return True
            else:
                stable_count = 0
                last_size = current_size
                
        except Exception as e:
            logger.debug(f"File not ready: {e}")
            stable_count = 0
            
        time.sleep(check_interval)
    
    return False


def prepare_file_for_com_access(file_path: str) -> None:
    """
    Prepare file for COM access by ensuring it's fully written and released.
    
    Args:
        file_path: Path to the file to prepare
        
    Raises:
        FileNotFoundError: If file doesn't exist
    """
    import os
    import time
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    logger.debug(f"Preparing file for COM access: {file_path}")
    
    # Force file system cache flush
    try:
        with open(file_path, 'rb+') as f:
            f.flush()
            os.fsync(f.fileno())
        logger.debug("File system cache flushed")
    except Exception as e:
        logger.debug(f"Could not flush file cache: {e}")
    
    # Small delay for file system
    time.sleep(0.5)
    
    # Verify file is accessible
    if not wait_for_file_ready(file_path, max_wait=5.0):
        logger.warning(f"File may not be fully ready: {file_path}")


@safe_excel_operation
def populate_do_tab_4_review_sheet_v2(processor: 'ExcelProcessor', input_wb: Any, dataframe_path: str) -> None:
    """
    Populate DO Tab 4 Review sheet using ExcelProcessor with best practices.
    
    Args:
        processor: ExcelProcessor instance
        input_wb: Input workbook COM object
        dataframe_path: Path to the dataframe Excel file
    """
    logger.info("Populating DO Tab 4 Review sheet with ExcelProcessor")
    
    if not os.path.exists(dataframe_path):
        logger.error(f"Dataframe file not found: {dataframe_path}")
        return
    
    try:
        # Open dataframe workbook
        df_wb = processor.open_workbook(dataframe_path, read_only=True)
        
        if df_wb:
            df_sheet = df_wb.Sheets(1)
            target_sheet = input_wb.Sheets("DO Tab 4 Review")
            
            # Clear existing data
            target_sheet.UsedRange.ClearContents()
            
            # Get all values from source sheet
            source_values = get_used_range_values(df_sheet)
            
            if source_values:
                # Write values to target sheet
                rows = len(source_values)
                cols = len(source_values[0]) if source_values else 0
                
                # Write data in chunks for better performance
                for row_idx, row_data in enumerate(source_values, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        target_sheet.Cells(row_idx, col_idx).Value = value
                
                # Apply date formatting
                apply_date_formatting(target_sheet)
                
                # Auto-fit columns
                target_sheet.UsedRange.Columns.AutoFit()
                
                logger.info(f"DO Tab 4 Review sheet populated with {rows} rows and {cols} columns")
            
    except Exception as e:
        logger.error(f"Error populating DO Tab 4 Review sheet: {e}")
        # Fall back to pandas method
        logger.info("Falling back to pandas method...")
        populate_do_tab_4_review_sheet_pandas(input_wb, dataframe_path)


def populate_do_tab_4_review_sheet(excel, input_wb, dataframe_path: str) -> None:
    """
    Populate the DO Tab 4 Review sheet with processed data.
    
    Args:
        excel: Excel application object
        input_wb: Input workbook object
        dataframe_path: Path to the Excel file containing processed dataframe
    """
    logger.info("Populating DO Tab 4 Review sheet with processed data")
    
    try:
        # Check if DO Tab 4 Review sheet exists
        sheet_exists = False
        for sheet in input_wb.Sheets:
            if sheet.Name == "DO Tab 4 Review":
                sheet_exists = True
                break
        
        if not sheet_exists:
            logger.error("DO Tab 4 Review sheet not found in workbook")
            return
        
        # Ensure the path is absolute and properly formatted
        dataframe_path = os.path.abspath(dataframe_path)
        
        # Load the processed data
        if not os.path.exists(dataframe_path):
            logger.error(f"Processed data file not found: {dataframe_path}")
            return
        
        logger.info(f"Opening dataframe from: {dataframe_path}")
        
        # Use enhanced robust method to open the dataframe workbook
        df_wb = None
        
        try:
            # Prepare file for COM access
            prepare_file_for_com_access(dataframe_path)
            
            # Small delay to ensure file system sync
            import time
            time.sleep(0.5)
            
            # Open with enhanced robust method
            df_wb = open_workbook_robust_v2(excel, dataframe_path, max_retries=5, read_only=True)
            logger.info("Successfully opened dataframe workbook with enhanced robust method")
            
        except Exception as e:
            logger.error(f"Failed to open dataframe file with enhanced robust method: {e}")
            # Fall back to pandas method
            logger.info("Falling back to pandas method...")
            import pandas as pd
            df = pd.read_excel(dataframe_path, sheet_name=0, engine='openpyxl')
            
            # Get the DO Tab 4 Review sheet
            target_sheet = input_wb.Sheets("DO Tab 4 Review")
            
            # Clear existing data
            target_sheet.UsedRange.ClearContents()
            
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                target_sheet.Cells(1, col_idx).Value = str(col_name)
            
            # Write data
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    if pd.notna(value):
                        if isinstance(value, pd.Timestamp):
                            target_sheet.Cells(row_idx + 2, col_idx).Value = value.to_pydatetime()
                        else:
                            target_sheet.Cells(row_idx + 2, col_idx).Value = value
            
            # Apply date formatting
            date_columns = [
                'Date of Advance', 'Last Activity Date', 'Anticipated Liquidation Date',
                'Period of Performance End Date', 'Date of Advance_comp', 
                'Last Activity Date_comp', 'Anticipated Liquidation Date_comp',
                'Period of Performance End Date_comp'
            ]
            
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in date_columns:
                    col_letter = get_column_letter(col_idx)
                    last_row = len(df) + 1
                    date_range = target_sheet.Range(f"{col_letter}2:{col_letter}{last_row}")
                    date_range.NumberFormat = "*m/dd/yyyy"
            
            logger.info("DO Tab 4 Review sheet populated successfully using pandas fallback")
            return
        
        # If we successfully opened with COM, continue with COM method
        if df_wb:
            df_sheet = df_wb.Sheets(1)
            
            # Get the DO Tab 4 Review sheet
            target_sheet = input_wb.Sheets("DO Tab 4 Review")
            
            # Clear existing data
            target_sheet.UsedRange.ClearContents()
            
            # Copy all data including headers
            df_sheet.UsedRange.Copy()
            target_sheet.Range("A1").PasteSpecial(-4163)  # xlPasteValues
            
            # Clear clipboard
            excel.CutCopyMode = False
            
            # Apply date formatting
            date_columns = [
                'Date of Advance', 'Last Activity Date', 'Anticipated Liquidation Date',
                'Period of Performance End Date', 'Date of Advance_comp', 
                'Last Activity Date_comp', 'Anticipated Liquidation Date_comp',
                'Period of Performance End Date_comp'
            ]
            
            # Find date columns and apply formatting
            header_row = 1
            for col in range(1, target_sheet.UsedRange.Columns.Count + 1):
                header_value = target_sheet.Cells(header_row, col).Value
                if header_value in date_columns:
                    # Apply date format to entire column
                    col_letter = get_column_letter(col)
                    last_row = target_sheet.UsedRange.Rows.Count
                    date_range = target_sheet.Range(f"{col_letter}2:{col_letter}{last_row}")
                    date_range.NumberFormat = "*m/dd/yyyy"
            
            # Auto-fit columns for better visibility
            target_sheet.UsedRange.Columns.AutoFit()
            
            # Close the dataframe workbook without saving
            df_wb.Close(SaveChanges=False)
            
            logger.info("DO Tab 4 Review sheet populated successfully using Excel COM")
            
            # Log sample of populated data
            logger.info("Sample of DO Tab 4 Review data (first 3 rows):")
            for row in range(1, min(4, target_sheet.UsedRange.Rows.Count + 1)):
                row_data = []
                for col in range(1, min(6, target_sheet.UsedRange.Columns.Count + 1)):
                    value = target_sheet.Cells(row, col).Value
                    if value is not None:
                        row_data.append(str(value)[:30])  # Limit string length for logging
                logger.info(f"Row {row}: {' | '.join(row_data)}")
        
    except Exception as e:
        logger.error(f"Error populating DO Tab 4 Review sheet: {str(e)}")
        raise


@safe_excel_operation
def process_excel_files_v2(output_path: str, input_path: str, current_dhstier_path: str, prior_dhstier_path: str, component: str, password: str, dataframe_path: str = None) -> None:
    """
    Process Excel files using the new ExcelProcessor with best practices.
    
    This is the recommended approach following Excel COM best practices.
    """
    if not EXCEL_PROCESSOR_AVAILABLE:
        logger.warning("ExcelProcessor not available, falling back to legacy method")
        return process_excel_files_legacy(output_path, input_path, current_dhstier_path, prior_dhstier_path, component, password, dataframe_path)
    
    logger.info(f"Starting Excel file processing with ExcelProcessor for {component}")
    
    # Use ExcelProcessor context manager for proper lifecycle management
    with ExcelProcessor() as processor:
        try:
            # Open all workbooks
            output_wb = processor.open_workbook(output_path)
            input_wb = processor.open_workbook(input_path)
            current_dhstier_wb = processor.open_workbook(current_dhstier_path)
            prior_dhstier_wb = processor.open_workbook(prior_dhstier_path)
            
            # Log sheet names for debugging
            logger.info("Sheet names in input workbook:")
            for sheet in input_wb.Sheets:
                logger.info(f"  - {sheet.Name}")
            
            # Find target sheet for insertion
            target_sheet_name = "6-ADVANCE TO TIER Recon Summary"
            target_sheet = processor.find_sheet(input_wb, target_sheet_name)
            
            if not target_sheet:
                logger.warning(f"Target sheet '{target_sheet_name}' not found")
                target_sheet = input_wb.Sheets(input_wb.Sheets.Count)  # Use last sheet
            
            # Copy DO Tab 4 Review sheet
            logger.info("Copying 'DO Tab 4 Review' sheet")
            source_sheet = processor.find_sheet(output_wb, "DO Tab 4 Review")
            if source_sheet:
                processor.copy_sheet(source_sheet, input_wb, after_sheet=target_sheet)
                target_sheet = input_wb.Sheets("DO Tab 4 Review")
            
            # Find and copy DHSTIER sheets
            logger.info("Finding and copying DHSTIER sheets")
            
            # Current year
            current_sheet = find_sheet_with_component_total(current_dhstier_wb, component)
            if current_sheet:
                sheet = current_dhstier_wb.Sheets(current_sheet)
                processor.copy_sheet(sheet, input_wb, "DO CY TB", after_sheet=target_sheet)
                target_sheet = input_wb.Sheets("DO CY TB")
            
            # Prior year
            prior_sheet = find_sheet_with_component_total(prior_dhstier_wb, component)
            if prior_sheet:
                sheet = prior_dhstier_wb.Sheets(prior_sheet)
                processor.copy_sheet(sheet, input_wb, "DO PY TB", after_sheet=target_sheet)
            
            # Process other operations (pivot tables, tickmarks, etc.)
            if password:
                # Use protected sheet operations
                logger.info("Processing protected sheets")
                for sheet in input_wb.Sheets:
                    with processor.protected_sheet_operation(sheet, password):
                        # Perform operations on unprotected sheet
                        pass
            
            # Populate DO Tab 4 Review if dataframe path provided
            if dataframe_path:
                logger.info("Populating DO Tab 4 Review sheet")
                populate_do_tab_4_review_sheet_v2(processor, input_wb, dataframe_path)
            
            # Save the workbook
            processor.save_workbook(input_wb)
            
            logger.info("Excel file processing completed successfully")
            
        except Exception as e:
            logger.error(f"Error in Excel file processing: {e}")
            raise


@safe_excel_operation
def process_excel_files(output_path: str, input_path: str, current_dhstier_path: str, prior_dhstier_path: str, component: str, password: str, dataframe_path: str = None) -> None:
    """
    Legacy process Excel files function.
    
    This function is maintained for backward compatibility but should be replaced
    with process_excel_files_v2 which uses the ExcelProcessor with best practices.
    """
    # Try to use the new version first
    if EXCEL_PROCESSOR_AVAILABLE:
        try:
            return process_excel_files_v2(output_path, input_path, current_dhstier_path, prior_dhstier_path, component, password, dataframe_path)
        except Exception as e:
            logger.warning(f"ExcelProcessor method failed, falling back to legacy: {e}")
    
    # Continue with existing implementation
    return process_excel_files_legacy(output_path, input_path, current_dhstier_path, prior_dhstier_path, component, password, dataframe_path)


@safe_excel_operation
def process_excel_files_legacy(output_path: str, input_path: str, current_dhstier_path: str, prior_dhstier_path: str, component: str, password: str, dataframe_path: str = None) -> None:
    """
    Process Excel files by copying sheets, creating pivot table, modifying sheets, and ensuring file accessibility.
    
    Args:
        output_path (str): Path to the source Excel file.
        input_path (str): Path to the destination Excel file.
        current_dhstier_path (str): Path to the current year DHSTIER Trial Balance file.
        prior_dhstier_path (str): Path to the prior year DHSTIER Trial Balance file.
        component (str): The component name (e.g., "WMD", "CBP").
        password (str): Password to unprotect sheets
        
    Raises:
        ValueError: If required sheets are not found.
        FileNotFoundError: If any of the input files don't exist.
        Exception: For other errors during processing.
    """
    logger.info(f"Starting Excel file processing for {component}")
    
    excel = None
    output_wb = None
    input_wb = None
    current_dhstier_wb = None
    prior_dhstier_wb = None
    
    try:
        # Check if files exist before opening
        logger.info("Validating file paths...")
        for file_path in [output_path, input_path, current_dhstier_path, prior_dhstier_path]:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
        
        # Prepare all files for COM access
        logger.info("Preparing files for COM access...")
        for file_path in [output_path, input_path, current_dhstier_path, prior_dhstier_path]:
            prepare_file_for_com_access(file_path)
        
        # Add delay between file operations to ensure file system sync
        import time
        time.sleep(1)

        # Initialize Excel COM with validation
        excel = initialize_excel_com(max_retries=3)
        
        # Open workbooks with enhanced robust strategy
        logger.info("Opening workbooks with enhanced strategy...")
        output_wb = open_workbook_robust_v2(excel, output_path, max_retries=3)
        
        # Add small delay before opening next workbook
        time.sleep(0.5)
        
        input_wb = open_workbook_robust_v2(excel, input_path, max_retries=3)
        
        # Add small delay before opening next workbook
        time.sleep(0.5)
        
        current_dhstier_wb = open_workbook_robust_v2(excel, current_dhstier_path, max_retries=3)
        
        # Add small delay before opening next workbook
        time.sleep(0.5)
        
        prior_dhstier_wb = open_workbook_robust_v2(excel, prior_dhstier_path, max_retries=3)
        
        # Log all sheet names in the input workbook for debugging
        logger.info("Sheet names in input workbook:")
        for sheet in input_wb.Sheets:
            logger.info(f"  - {sheet.Name}")

        # Find the "6-ADVANCE TO TIER Recon Summary" sheet
        target_sheet = "6-ADVANCE TO TIER Recon Summary"
        sheet_found = False
        
        # First try exact match
        for sheet in input_wb.Sheets:
            if sheet.Name == target_sheet:
                sheet_found = True
                break
        
        # If not found, try case-insensitive search and partial match
        if not sheet_found:
            logger.warning(f"Sheet '{target_sheet}' not found with exact match. Trying case-insensitive search...")
            for sheet in input_wb.Sheets:
                if "advance" in sheet.Name.lower() and "tier" in sheet.Name.lower() and "recon" in sheet.Name.lower():
                    logger.info(f"Found similar sheet: '{sheet.Name}'. Using this instead.")
                    target_sheet = sheet.Name
                    sheet_found = True
                    break
                
        if not sheet_found:
            logger.error(f"Sheet '{target_sheet}' not found in the input workbook")
            logger.error("Available sheets are:")
            for sheet in input_wb.Sheets:
                logger.error(f"  - {sheet.Name}")
            # Continue processing without this sheet rather than failing
            logger.warning("Continuing without the ADVANCE TO TIER Recon Summary sheet")
            target_sheet = None

        # Copy "DO Tab 4 Review" sheet
        if target_sheet:
            logger.info("Copying 'DO Tab 4 Review' sheet")
            advanced_copy_sheet(output_wb, input_wb, "DO Tab 4 Review", insert_after=target_sheet)
            target_sheet = "DO Tab 4 Review"  # Update target_sheet for next insertion
        else:
            logger.warning("Skipping 'DO Tab 4 Review' sheet copy due to missing target sheet")

        # Copy and rename current year DHSTIER sheet
        if target_sheet:
            logger.info("Copying current year DHSTIER sheet")
            current_sheet_name = find_sheet_name(current_dhstier_wb, component)
            advanced_copy_sheet(current_dhstier_wb, input_wb, current_sheet_name, "DO CY TB", insert_after=target_sheet)
            target_sheet = "DO CY TB"  # Update target_sheet for next insertion
        else:
            logger.warning("Skipping current year DHSTIER sheet copy due to missing target sheet")

        # Copy and rename prior year DHSTIER sheet
        if target_sheet:
            logger.info("Copying prior year DHSTIER sheet")
            prior_sheet_name = find_sheet_name(prior_dhstier_wb, component)
            advanced_copy_sheet(prior_dhstier_wb, input_wb, prior_sheet_name, "DO PY TB", insert_after=target_sheet)
        else:
            logger.warning("Skipping prior year DHSTIER sheet copy due to missing target sheet")
        
        # Populate DO Tab 4 Review sheet with processed data if available
        if dataframe_path:
            # Ensure absolute path
            dataframe_path = os.path.abspath(dataframe_path)
            
            if os.path.exists(dataframe_path):
                # Small delay to ensure file is not locked
                import time
                time.sleep(0.5)
                
                # Log file size to verify it's fully written
                file_size = os.path.getsize(dataframe_path)
                logger.info(f"DO Tab 4 Review data file size: {file_size} bytes")
                
                populate_do_tab_4_review_sheet(excel, input_wb, dataframe_path)
            else:
                logger.warning(f"Processed dataframe file not found: {dataframe_path}")
        else:
            logger.warning("No processed dataframe path provided, skipping DO Tab 4 Review population")
        
        # Save after copying all sheets
        try:
            input_wb.Save()
            logger.info("Saved workbook after copying sheets")
        except Exception as e:
            logger.warning(f"Could not save after copying sheets: {str(e)}")

        # Create pivot table and get sum_cell_address
        sum_cell_address = create_pivot_table(input_wb, password)
        logger.info(f"Sum cell address: {sum_cell_address}")
        
        # Save after pivot table creation
        try:
            input_wb.Save()
            logger.info("Saved workbook after pivot table creation")
        except Exception as e:
            logger.warning(f"Could not save after pivot table creation: {str(e)}")

        # Create tickmark legend and compare values
        logger.info("Creating tickmark legend and comparing values")
        create_tickmark_legend_and_compare_values(input_wb, password)
        
        # Save after tickmark creation
        try:
            input_wb.Save()
            logger.info("Saved workbook after tickmark creation")
        except Exception as e:
            logger.warning(f"Could not save after tickmark creation: {str(e)}")

        # Modify the Obligation Analysis sheet and get last_column, header_row, and sum_udo_balance_col2 for table comparison
        logger.info("Modifying Obligation Analysis sheet")
        last_column, header_row, sum_udo_balance_col2 = modify_obligation_analysis_sheet(input_wb, password, component)
        # Note: sum_udo_balance_col2 is reserved for future use in validate_udo_tier_recon function
        
        # Compare Obligation Analysis tables and add tickmarks
        logger.info("Comparing Obligation Analysis tables and adding tickmarks")
        compare_obligation_analysis_tables(input_wb, password, last_column, header_row)

        # Add UDO TIER reconciliation validation
        # TODO: Implement validate_udo_tier_recon function or find the missing module
        logger.info("Skipping UDO TIER reconciliation validation - function not implemented")
        # validate_udo_tier_recon(excel, input_wb, component, password, sum_cell_address, sum_udo_balance_col2)
        # logger.info("UDO TIER reconciliation validation completed")

        # Save and close workbooks
        input_wb.Save()
        logger.info(f"Excel file processing completed successfully for {component}")

    except FileNotFoundError as e:
        logger.error(f"File not found: {str(e)}")
        raise
    except ValueError as e:
        logger.error(f"Value error in Excel processing: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Error in Excel file processing: {str(e)}", exc_info=True)
        raise
    finally:
        # Save the input workbook if it was modified
        if input_wb:
            try:
                logger.info("Saving input workbook before closing...")
                input_wb.Save()
                logger.info("Input workbook saved successfully")
            except Exception as save_error:
                logger.error(f"Error saving input workbook: {str(save_error)}")
        
        # Use enhanced cleanup process
        workbooks = [output_wb, current_dhstier_wb, prior_dhstier_wb, input_wb]
        cleanup_com_objects(excel, workbooks)

    # Ensure file accessibility after processing
    try:
        ensure_file_accessibility(input_path)
        logger.info(f"File accessibility ensured for {input_path}")
    except Exception as e:
        logger.error(f"Error ensuring file accessibility: {str(e)}", exc_info=True)
        raise


@safe_excel_operation
def find_column(sheet, header_row: int, column_name: str) -> Optional[int]:
    """
    Finds the column number for a given column name.
    
    Args:
        sheet: Excel worksheet object
        header_row (int): Row number containing headers
        column_name (str): The header text to search for
    
    Returns:
        Optional[int]: The column index if found, None otherwise
    """
    for col in range(1, sheet.UsedRange.Columns.Count + 1):
        if sheet.Cells(header_row, col).Value == column_name:
            return col
    logger.warning(f"Column '{column_name}' not found")
    return None


@safe_excel_operation
def add_count_formula(sheet, header_row: int, last_row: int, count_col: int, active_col: int, 
                      status_col: int, criteria_col: int, criteria: str) -> None:
    """
    Adds the Count formula and formatting for Active, Inactive, and No Invoice Activity.
    
    Args:
        sheet: Excel worksheet object
        header_row (int): Row number containing headers
        last_row (int): Last row number with data
        count_col (int): Column where count formula will be added
        active_col (int): Column containing active/inactive status
        status_col (int): Column containing status values
        criteria_col (int): Column containing criteria values
        criteria (str): Criteria string for the COUNTIFS formula
    """
    for i in range(5):
        cell = sheet.Cells(header_row + 1 + i, count_col)
        if i < 4:
            cell.Formula = f'=COUNTIFS({get_column_letter(active_col)}${header_row + 1}:{get_column_letter(active_col)}${last_row},"{criteria}",{get_column_letter(status_col)}${header_row + 1}:{get_column_letter(status_col)}${last_row},{get_column_letter(criteria_col)}{header_row + 1 + i})'
        else:
            cell.Formula = f'=SUM({get_column_letter(count_col)}{header_row + 1}:{get_column_letter(count_col)}{header_row + 4})'
        cell.Font.Bold = True
        cell.Font.Name = "Calibri"
        cell.Font.Size = 11
        cell.NumberFormat = "#,##0_);(#,##0)"


@safe_excel_operation
def find_column_index(sheet, header_text: str, row: int) -> Optional[int]:
    """
    Finds the column index for a given header text in the specified row of the sheet.
    
    Args:
        sheet: Excel worksheet object
        header_text (str): The header text to search for
        row (int): The row number to search in
    
    Returns:
        Optional[int]: The column index if found, None otherwise
    """
    try:
        for cell in sheet.Range(f"{row}:{row}"):
            if cell.Value and header_text.lower() in str(cell.Value).lower():
                return cell.Column
        logger.warning(f"Header '{header_text}' not found in the sheet at row {row}.")
        return None
    except Exception as e:
        logger.error(f"Error finding column index for '{header_text}': {str(e)}", exc_info=True)
        return None


@safe_excel_operation
def find_header_row(sheet) -> Optional[int]:
    """
    Finds the header row in the given sheet.
    
    Args:
        sheet: Excel worksheet object
    
    Returns:
        Optional[int]: The row number of the header if found, None otherwise
    """
    try:
        for row in range(1, 20):  # Adjust the range as needed
            cell = sheet.Cells(row, 1)
            if cell.Value and "TAS" in str(cell.Value):
                return row
        logger.warning("Header row not found in the sheet.")
        return None
    except Exception as e:
        logger.error(f"Error finding header row: {str(e)}", exc_info=True)
        return None


@safe_excel_operation
def find_keyword_column(sheet, header_row: int, keywords: List[str]) -> Optional[int]:
    """
    Find the first column that matches any of the given keywords.
    
    Args:
        sheet: Excel worksheet object
        header_row (int): Row number containing headers
        keywords (List[str]): List of keywords to search for
    
    Returns:
        Optional[int]: Column index if found, None otherwise
    """
    for col in range(1, sheet.UsedRange.Columns.Count + 1):
        cell_value = str(sheet.Cells(header_row, col).Value).lower()
        if any(keyword.lower() in cell_value for keyword in keywords):
            return col
    return None


@safe_excel_operation
def modify_obligation_analysis_sheet(wb, password: str, component: str) -> Tuple[int, int, int]:
    """
    Modifies the Obligation Analysis sheet with advanced modifications including new columns, 
    formulas, and formatting.
    
    Args:
        wb: Workbook (win32com.client.CDispatch object)
        password (str): Password to unprotect sheets
        component (str): The component selected by the user
        
    Returns:
        Tuple[int, int, int]: The last column, header row, and the column index of the second "Sum of UDO Balance"
    """
    try:
        # Find the correct sheets
        target_sheet = None
        do_tab_4_review_sheet = None
        for sheet in wb.Sheets:
            if "Obligation Analysis" in sheet.Name or "4-Obligation Analysis" in sheet.Name:
                target_sheet = sheet
            elif "DO Tab 4 Review" in sheet.Name:
                do_tab_4_review_sheet = sheet
        
        if not target_sheet or not do_tab_4_review_sheet:
            raise ValueError("Required sheets not found")
        
        logger.info(f"Modifying sheet: {target_sheet.Name}")
        
        # Unprotect the sheets
        target_sheet.Unprotect(Password=password)
        do_tab_4_review_sheet.Unprotect(Password=password)
        
        # Find the "TAS" cell in column A
        tas_cell = target_sheet.Cells.Find("TAS", After=target_sheet.Cells(1, 1), 
                                          LookIn=win32com.client.constants.xlValues, 
                                          LookAt=win32com.client.constants.xlWhole)
        if not tas_cell:
            raise ValueError("TAS cell not found in column A")
        
        logger.info(f"TAS cell found at row {tas_cell.Row}, column {tas_cell.Column}")
        
        header_row = tas_cell.Row
        
        # Find the last column
        last_column = target_sheet.Cells(header_row, target_sheet.Columns.Count).End(win32com.client.constants.xlToLeft).Column
        logger.info(f"Last populated column in header row: {get_column_letter(last_column)}")
        
        # Log all header row values for debugging
        logger.info(f"Header row values in '{target_sheet.Name}' sheet:")
        for col in range(1, min(last_column + 1, 20)):  # Log first 20 columns to avoid too much output
            cell_value = target_sheet.Cells(header_row, col).Value
            col_letter = get_column_letter(col)
            logger.info(f"  Column {col_letter} (index {col}): {cell_value}")
        if last_column > 20:
            logger.info(f"  ... and {last_column - 20} more columns")

        # Find the header row in DO Tab 4 Review sheet
        do_tab_header_row = find_header_row(do_tab_4_review_sheet)

        if do_tab_header_row is None:
            raise ValueError("Header row not found in DO Tab 4 Review sheet")

        # Find the DO Concatenate column index in DO Tab 4 Review sheet
        do_concatenate_review_col = find_column_index(do_tab_4_review_sheet, "DO Concatenate", do_tab_header_row)

        if do_concatenate_review_col is None:
            raise ValueError("DO Concatenate column not found in DO Tab 4 Review sheet")

        logger.info(f"DO Concatenate column found in DO Tab 4 Review sheet: {do_concatenate_review_col}")

        # Find column indices relative to DO Concatenate column
        udo_age_col = find_column_index(do_tab_4_review_sheet, "UDO by Age", do_tab_header_row) - do_concatenate_review_col + 1
        active_inactive_col = find_column_index(do_tab_4_review_sheet, "Active / Inactive Obligation (No Invoice in Over 1 Year)", do_tab_header_row) - do_concatenate_review_col + 1
        do_comment_col = find_column_index(do_tab_4_review_sheet, "DO Comment", do_tab_header_row) - do_concatenate_review_col + 1

        if not all([udo_age_col > 0, active_inactive_col > 0, do_comment_col > 0]):
            raise ValueError("One or more required columns not found after DO Concatenate column in DO Tab 4 Review sheet")

        logger.info(f"Relative column indices found in DO Tab 4 Review sheet - UDO by Age: {udo_age_col}, Active/Inactive: {active_inactive_col}, DO Comment: {do_comment_col}")
        
        # Add new column headers
        new_headers = [
            "DO Concatenate", "UDO By Age Group", "Active / Inactive Obligation (No Invoice in Over 1 Year)", "DO Comment",
            "", "", "", "", "",  # 5 blank columns
            "UDO by Age Group", "Sum of UDO Balance", "UDO %",
            "", "",  # 2 blank columns
            "Status", "Sum of UDO Balance", "$ UDO as a % of Total", "Count of Active Obligations",
            "Count of Inactive Obligations", "Count of Obligations Without Invoice Activity"
        ]
        
        # Add headers and set initial column widths
        for i, header in enumerate(new_headers):
            col = last_column + i + 1
            cell = target_sheet.Cells(header_row, col)
            cell.Value = header
            if header:  # Only format non-empty cells
                cell.Interior.Color = 65535  # Yellow
                cell.Font.Color = 255  # Red
                cell.Font.Bold = True
                cell.Font.Name = "Calibri"
                cell.Font.Size = 11
                cell.HorizontalAlignment = win32com.client.constants.xlCenter
                cell.VerticalAlignment = win32com.client.constants.xlCenter
                cell.Borders.Weight = win32com.client.constants.xlThick
                cell.WrapText = True

                # Set specific column widths for the first 4 new columns
                if i == 0:
                    target_sheet.Columns(col).ColumnWidth = 70  # 1st new column
                elif i == 1:
                    target_sheet.Columns(col).ColumnWidth = 30  # 2nd new column
                elif i in [2, 3]:
                    target_sheet.Columns(col).ColumnWidth = 85  # 3rd and 4th new columns
                else:
                    target_sheet.Columns(col).ColumnWidth = 30  # Default width for other new columns
            else:
                target_sheet.Columns(col).ColumnWidth = 10  # Smaller width for blank headers

        # Find the last row
        last_row = target_sheet.Cells(target_sheet.Rows.Count, 1).End(win32com.client.constants.xlUp).Row
        logger.info(f"Last populated row: {last_row}")

        # Apply formatting to data rows for the first 4 new columns
        for i in range(4):
            col = last_column + i + 1
            range_to_format = target_sheet.Range(target_sheet.Cells(header_row + 1, col), target_sheet.Cells(last_row, col))
            range_to_format.Font.Color = 255  # Red
            range_to_format.Font.Bold = True
            range_to_format.Font.Name = "Calibri"
            range_to_format.Font.Size = 11

        # Apply wrap text to specific columns and set column width
        for col in [last_column + 3, last_column + 4]:  # "Active / Inactive Obligation" and "DO Comment" columns
            range_to_format = target_sheet.Range(target_sheet.Cells(header_row + 1, col), target_sheet.Cells(last_row, col))
            range_to_format.WrapText = True
            range_to_format.Font.Color = 255  # Red
            range_to_format.Font.Bold = True
            range_to_format.Font.Name = "Calibri"
            range_to_format.Font.Size = 11

        # Define column positions for formulas
        do_concatenate_col = last_column + 1
        udo_age_formula_col = last_column + 2
        active_inactive_formula_col = last_column + 3
        do_comment_formula_col = last_column + 4

        # Insert the DO Concatenate formula
        first_data_row = header_row + 1

        # Find column indices in target sheet
        other_identifier_col = find_column_index(target_sheet, "Other Unique Identifier if DHS Doc No is not unique1", header_row)

        # Find the first keyword column
        keyword_columns = ["PONO", "Item", "Line", "MDL"]  # Add more keywords if needed
        keyword_col = find_keyword_column(target_sheet, header_row, keyword_columns)

        SPECIAL_FORMULA_COMPONENTS = ["SS", "CBP", "MGA", "OIG", "FEM"]

        if not keyword_col:
            logger.warning("No keyword column found. Using a default formula.")
            do_concatenate_formula = f'=CONCATENATE(TRIM(A{first_data_row}),TRIM(C{first_data_row}),TRIM(D{first_data_row}))'
        else:
            logger.info(f"Keyword column found: {get_column_letter(keyword_col)}")
            if other_identifier_col:
                do_concatenate_formula = (
                    f'=CONCATENATE(TRIM(A{first_data_row}),TRIM(C{first_data_row}),'
                    f'IF(ISBLANK({get_column_letter(other_identifier_col)}{first_data_row}),'
                    f'TRIM({get_column_letter(keyword_col)}{first_data_row}),'
                    f'TRIM({get_column_letter(other_identifier_col)}{first_data_row})))'
                )
            else:
                do_concatenate_formula = f'=CONCATENATE(TRIM(A{first_data_row}),TRIM(C{first_data_row}),TRIM({get_column_letter(keyword_col)}{first_data_row}))'

        if component in SPECIAL_FORMULA_COMPONENTS:
            logger.info(f"Applying special DO Concatenate formula for component: {component}")
            do_concatenate_formula = (
                f'=CONCATENATE(TRIM(A{first_data_row}),TRIM(C{first_data_row}),'
                f'IF(MOD(ROUND(E{first_data_row},2),1)=0,'
                f'TEXT(ROUND(E{first_data_row},2),"0"),'
                f'IF(RIGHT(TEXT(ROUND(E{first_data_row},2),"0.00"),1)="0",'
                f'LEFT(TEXT(ROUND(E{first_data_row},2),"0.00"),LEN(TEXT(ROUND(E{first_data_row},2),"0.00"))-1),'
                f'TEXT(ROUND(E{first_data_row},2),"0.00"))))'
            )

        formula_range = target_sheet.Range(target_sheet.Cells(first_data_row, do_concatenate_col), 
                                          target_sheet.Cells(last_row, do_concatenate_col))
        formula_range.Formula = do_concatenate_formula

        logger.info(f"DO Concatenate formula for component {component} inserted and filled down from row {first_data_row} to {last_row}")

        # Insert UDO by Age formula
        udo_age_formula = f'=VLOOKUP({get_column_letter(do_concatenate_col)}{first_data_row},\'DO Tab 4 Review\'!{get_column_letter(do_concatenate_review_col)}:{get_column_letter(do_tab_4_review_sheet.UsedRange.Columns.Count)},{udo_age_col},FALSE)'
        formula_range = target_sheet.Range(target_sheet.Cells(first_data_row, udo_age_formula_col), 
                                          target_sheet.Cells(last_row, udo_age_formula_col))
        formula_range.Formula = udo_age_formula

        # Insert Active/Inactive Obligation formula
        active_inactive_formula = f'=VLOOKUP({get_column_letter(do_concatenate_col)}{first_data_row},\'DO Tab 4 Review\'!{get_column_letter(do_concatenate_review_col)}:{get_column_letter(do_tab_4_review_sheet.UsedRange.Columns.Count)},{active_inactive_col},FALSE)'
        formula_range = target_sheet.Range(target_sheet.Cells(first_data_row, active_inactive_formula_col), 
                                          target_sheet.Cells(last_row, active_inactive_formula_col))
        formula_range.Formula = active_inactive_formula

        # Insert DO Comment formula
        do_comment_formula = f'=VLOOKUP({get_column_letter(do_concatenate_col)}{first_data_row},\'DO Tab 4 Review\'!{get_column_letter(do_concatenate_review_col)}:{get_column_letter(do_tab_4_review_sheet.UsedRange.Columns.Count)},{do_comment_col},FALSE)'
        formula_range = target_sheet.Range(target_sheet.Cells(first_data_row, do_comment_formula_col), 
                                          target_sheet.Cells(last_row, do_comment_formula_col))
        formula_range.Formula = do_comment_formula

        logger.info(f"Additional formulas inserted and filled down from row {first_data_row} to {last_row}")
        
        # Define column positions based on new_headers list
        do_concatenate_col = last_column + 1
        udo_by_age_group_col = last_column + 2
        active_inactive_col = last_column + 3
        do_comment_col = last_column + 4
        udo_age_group_col = last_column + 10  # This is the second "UDO by Age Group" column
        sum_udo_balance_col = last_column + 11
        udo_percent_col = last_column + 12

        # Add values under the second "UDO By Age Group" column
        age_group_values = ["1) <= 360 Days", "2) 361 - 720 Days", "3) 721 - 1,080 Days", "4) > 1080 Days", "Grand Total"]
        for i, value in enumerate(age_group_values):
            cell = target_sheet.Cells(header_row + 1 + i, udo_age_group_col)
            cell.Value = value
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"

        # Add SUMIF formula in "Sum of UDO Balance" column
        for i in range(4):
            cell = target_sheet.Cells(header_row + 1 + i, sum_udo_balance_col)
            cell.Formula = f'=SUMIF({get_column_letter(udo_by_age_group_col)}${first_data_row}:{get_column_letter(udo_by_age_group_col)}${last_row},{get_column_letter(udo_age_group_col)}{header_row + 1 + i},M${first_data_row}:M${last_row})'
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"
            cell.Font.Size = 11
            cell.NumberFormat = "$#,##0.00_);[Red]($#,##0.00)"

        # Add SUM formula for Grand Total
        grand_total_cell = target_sheet.Cells(header_row + 5, sum_udo_balance_col)
        grand_total_cell.Formula = f'=SUM({get_column_letter(sum_udo_balance_col)}{header_row + 1}:{get_column_letter(sum_udo_balance_col)}{header_row + 4})'
        grand_total_cell.Font.Bold = True
        grand_total_cell.Font.Name = "Calibri"
        grand_total_cell.Font.Size = 11
        grand_total_cell.NumberFormat = "$#,##0.00_);[Red]($#,##0.00)"

        # Add percentage formula in "UDO %" column
        for i in range(5):
            cell = target_sheet.Cells(header_row + 1 + i, udo_percent_col)
            if i < 4:
                cell.Formula = f'={get_column_letter(sum_udo_balance_col)}{header_row + 1 + i}/{get_column_letter(sum_udo_balance_col)}{header_row + 5}'
            else:
                cell.Formula = f'=SUM({get_column_letter(udo_percent_col)}{header_row + 1}:{get_column_letter(udo_percent_col)}{header_row + 4})'
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"
            cell.Font.Color = 0  # Black
            cell.NumberFormat = "0%"

        # Define additional column positions
        wingdings_col1 = last_column + 13
        status_col = last_column + 15
        sum_udo_balance_col2 = last_column + 16
        udo_percent_col2 = last_column + 17
        count_active_col = last_column + 18
        count_inactive_col = last_column + 19
        count_no_invoice_col = last_column + 20
        wingdings_col2 = last_column + 21

        # Apply Wingdings font to first Wingdings column
        logger.info("Applying Wingdings font to first Wingdings column")
        for i in range(5):
            cell = target_sheet.Cells(header_row + 1 + i, wingdings_col1)
            cell.Font.Name = "Wingdings"
            cell.Font.Size = 10

        # Add Status column values
        logger.info("Adding Status column values")
        status_values = ["1", "2", "3", "4", "Grand Total"]
        for i, value in enumerate(status_values):
            cell = target_sheet.Cells(header_row + 1 + i, status_col)
            cell.Value = value
            cell.Font.Color = 255  # Red
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"
            cell.HorizontalAlignment = win32com.client.constants.xlLeft

        # Add second "Sum of UDO Balance" column formula
        logger.info("Adding second Sum of UDO Balance column formula")
        current_quarter_status_col = find_column(target_sheet, header_row, "Current Quarter Status")
        current_fy_balance_col = find_column(target_sheet, header_row, "Current FY Quarter-End  balance UDO")
        for i in range(5):
            cell = target_sheet.Cells(header_row + 1 + i, sum_udo_balance_col2)
            if i < 4:
                cell.Formula = f'=SUMIF({get_column_letter(current_quarter_status_col)}${first_data_row}:{get_column_letter(current_quarter_status_col)}${last_row},{get_column_letter(status_col)}{header_row + 1 + i},{get_column_letter(current_fy_balance_col)}${first_data_row}:{get_column_letter(current_fy_balance_col)}${last_row})'
            else:
                cell.Formula = f'=SUM({get_column_letter(sum_udo_balance_col2)}{header_row + 1}:{get_column_letter(sum_udo_balance_col2)}{header_row + 4})'
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"
            cell.Font.Size = 11
            cell.NumberFormat = "$#,##0.00_);[Red]($#,##0.00)"

        # Add "$ UDO as a % of Total" column formula
        logger.info("Adding $ UDO as a % of Total column formula")
        for i in range(5):
            cell = target_sheet.Cells(header_row + 1 + i, udo_percent_col2)
            if i < 4:
                cell.Formula = f'={get_column_letter(sum_udo_balance_col2)}{header_row + 1 + i}/{get_column_letter(sum_udo_balance_col2)}{header_row + 5}'
            else:
                cell.Formula = f'=SUM({get_column_letter(udo_percent_col2)}{header_row + 1}:{get_column_letter(udo_percent_col2)}{header_row + 4})'
            cell.Font.Bold = True
            cell.Font.Name = "Calibri"
            cell.Font.Color = 0  # Black
            cell.NumberFormat = "0.00%"

        # Add Count formulas
        logger.info("Adding Count formulas")
        active_inactive_col = find_column(target_sheet, header_row, "Active / Inactive Obligation (No Invoice in Over 1 Year)")
        add_count_formula(target_sheet, header_row, last_row, count_active_col, active_inactive_col, current_quarter_status_col, status_col, "Active Obligation — Invoice Received in Last 12 Months")
        add_count_formula(target_sheet, header_row, last_row, count_inactive_col, active_inactive_col, current_quarter_status_col, status_col, "Inactive Obligation — No Invoice Activity Within Last 12 Months")
        add_count_formula(target_sheet, header_row, last_row, count_no_invoice_col, active_inactive_col, current_quarter_status_col, status_col, "No Invoice Activity Reported")

        # Add "Count of Total Obligations" label and formula
        logger.info("Adding Count of Total Obligations")
        label_cell = target_sheet.Cells(header_row + 7, count_inactive_col)
        label_cell.Value = "Count of Total Obligations"
        label_cell.Font.Color = 255  # Red
        label_cell.Font.Bold = True
        label_cell.Font.Name = "Calibri"
        label_cell.Font.Size = 11

        total_cell = target_sheet.Cells(header_row + 7, count_no_invoice_col)
        total_cell.Formula = f'=SUM({get_column_letter(count_active_col)}{header_row + 5}:{get_column_letter(count_no_invoice_col)}{header_row + 5})'
        total_cell.Font.Bold = True
        total_cell.Font.Name = "Calibri"
        total_cell.Font.Size = 11
        total_cell.NumberFormat = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

        # Apply Wingdings font to second Wingdings column
        logger.info("Applying Wingdings font to second Wingdings column")
        for i in range(7):
            cell = target_sheet.Cells(header_row + 1 + i, wingdings_col2)
            cell.Font.Name = "Wingdings"
            cell.Font.Size = 10

        # --- Robustly find the second "Sum of UDO Balance" column ---
        # This is the column in the new summary table (not the first one in the age group table)
        header_row_cells = target_sheet.Range(f"{header_row}:{header_row}")
        sum_udo_balance_cols = []
        for cell in header_row_cells:
            if cell.Value and "sum of udo balance" in str(cell.Value).lower():
                sum_udo_balance_cols.append(cell.Column)
        if len(sum_udo_balance_cols) < 2:
            logger.error("Could not find the second 'Sum of UDO Balance' column in Obligation Analysis sheet.")
            sum_udo_balance_col2 = None
        else:
            sum_udo_balance_col2 = sum_udo_balance_cols[1]
            logger.info(f"Second 'Sum of UDO Balance' column found at index: {sum_udo_balance_col2}")

        logger.info("Advanced modifications to Obligation Analysis sheet completed successfully")
        return last_column, header_row, sum_udo_balance_col2
    except Exception as e:
        logger.error(f"Error in advanced modifications to Obligation Analysis sheet: {str(e)}", exc_info=True)
        raise


@safe_excel_operation
def compare_obligation_analysis_tables(wb, password: str, last_column: int, header_row: int) -> None:
    """
    Compares data between Certification and Obligation Analysis sheets and adds tickmarks.
    
    Args:
        wb: Workbook (win32com.client.CDispatch object)
        password (str): Password to unprotect sheets
        last_column (int): Last column position from modify_obligation_analysis_sheet
        header_row (int): Header row position from modify_obligation_analysis_sheet
    """
    try:
        logger.info("Starting comparison of Obligation Analysis tables")

        cert_sheet = wb.Worksheets("2-Certification")
        obl_sheet = wb.Worksheets("4-Obligation Analysis")

        cert_sheet.Unprotect(Password=password)
        obl_sheet.Unprotect(Password=password)

        # First comparison: Obligation Analysis
        compare_table(cert_sheet, obl_sheet, "Obligation Analysis", last_column, header_row, 15, 20, 2, 5)

        # Second comparison: UDO Balance by Age
        compare_table(cert_sheet, obl_sheet, "UDO Balance by Age", last_column, header_row, 10, 12, 2, 3)

        logger.info("Comparison of Obligation Analysis tables completed successfully")

    except Exception as e:
        logger.error(f"Error in comparing Obligation Analysis tables: {str(e)}", exc_info=True)
        raise


@safe_excel_operation
def compare_table(cert_sheet, obl_sheet, table_name: str, last_column: int, header_row: int, 
                 start_col_offset: int, end_col_offset: int, 
                 start_compare_col: int, end_compare_col: int) -> None:
    """
    Compare a specific table between certification and obligation analysis sheets.
    
    Args:
        cert_sheet: Certification worksheet object
        obl_sheet: Obligation Analysis worksheet object
        table_name (str): Name of the table to compare
        last_column (int): Last column position from modify_obligation_analysis_sheet
        header_row (int): Header row position from modify_obligation_analysis_sheet
        start_col_offset (int): Starting column offset
        end_col_offset (int): Ending column offset
        start_compare_col (int): First column to compare
        end_compare_col (int): Last column to compare
    """
    try:
        logger.info(f"Starting comparison of {table_name} tables")

        # Find the instance of table_name in Certification sheet
        first_instance = cert_sheet.Cells.Find(table_name, 
                                              LookIn=win32com.client.constants.xlValues, 
                                              LookAt=win32com.client.constants.xlWhole)
        if not first_instance:
            raise ValueError(f"'{table_name}' not found in Certification sheet")

        # For Obligation Analysis, find the second instance
        if table_name == "Obligation Analysis":
            table_cell = cert_sheet.Cells.FindNext(first_instance)
            if not table_cell or table_cell.Address == first_instance.Address:
                raise ValueError("Second instance of 'Obligation Analysis' not found in Certification sheet")
        else:
            table_cell = first_instance

        logger.info(f"'{table_name}' found in Certification sheet at {table_cell.Address}")

        # Find last populated column in the same row
        last_col = cert_sheet.Cells(table_cell.Row, cert_sheet.Columns.Count).End(win32com.client.constants.xlToLeft).Column
        last_col_cell = cert_sheet.Cells(table_cell.Row, last_col)
        logger.info(f"Last populated column in Certification sheet at {last_col_cell.Address}")

        # Define the range for the Certification sheet table (excluding header row)
        cert_range = cert_sheet.Range(cert_sheet.Cells(table_cell.Row + 1, table_cell.Column), 
                                      cert_sheet.Cells(table_cell.Row + 5, last_col))
        logger.info(f"Certification sheet table range: {cert_range.Address}")

        # Define the range for the Obligation Analysis sheet table (excluding header row)
        start_col = last_column + start_col_offset
        end_col = last_column + end_col_offset
        obl_range = obl_sheet.Range(
            obl_sheet.Cells(header_row + 1, start_col),
            obl_sheet.Cells(header_row + 5, end_col)
        )
        logger.info(f"Obligation Analysis sheet table range: {obl_range.Address}")

        # Compare the tables
        for row in range(1, 6):  # 5 data rows
            cert_row = cert_range.Rows(row)
            obl_row = obl_range.Rows(row)

            logger.info(f"Comparing row {row}:")
            cert_values = [cert_row.Cells(1, col).Value for col in range(1, end_compare_col + 1)]
            obl_values = [obl_row.Cells(1, col).Value for col in range(1, end_compare_col + 1)]
            logger.info(f"  Certification: {cert_values}")
            logger.info(f"  Obligation Analysis: {obl_values}")

            values_match = True
            for col in range(start_compare_col, end_compare_col + 1):
                cert_value = cert_row.Cells(1, col).Value
                obl_value = obl_row.Cells(1, col).Value

                if cert_value != obl_value:
                    logger.warning(f"Mismatch found at row {row}, column {col}. Cert: {cert_value}, Obl: {obl_value}")
                    values_match = False
                    break

            if values_match:
                logger.info(f"Row {row} matches between tables")

                # Add tickmark to Certification sheet
                cert_tickmark_cell = cert_sheet.Cells(cert_row.Row, last_col + 1)
                cert_tickmark_cell.Font.Name = "Wingdings"
                cert_tickmark_cell.Font.Color = 0  # Black
                cert_tickmark_cell.Font.Size = 10
                cert_tickmark_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                cert_tickmark_cell.Value = "h"
                logger.info(f"Added 'h' tickmark to Certification sheet at {cert_tickmark_cell.Address}")

                # Add tickmark to Obligation Analysis sheet
                obl_tickmark_cell = obl_sheet.Cells(obl_row.Row, end_col + 1)
                obl_tickmark_cell.Font.Name = "Wingdings"
                obl_tickmark_cell.Font.Color = 0  # Black
                obl_tickmark_cell.Font.Size = 10
                obl_tickmark_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                obl_tickmark_cell.Value = "m"
                logger.info(f"Added 'm' tickmark to Obligation Analysis sheet at {obl_tickmark_cell.Address}")
            else:
                logger.warning(f"Row {row} does not match between tables")

        # Additional check for Obligation Analysis table
        if table_name == "Obligation Analysis":
            # Get the check figure from the Certification sheet
            check_row = cert_range.Row + cert_range.Rows.Count
            check_col = last_col
            check_figure = abs(cert_sheet.Cells(check_row, check_col).Value)
            formatted_check_figure = format_excel_style(check_figure)
            logger.info(f"Check figure at {cert_sheet.Cells(check_row, check_col).Address}: {formatted_check_figure}")

            # Get the total from the row above in the Certification sheet
            total_above = cert_sheet.Cells(check_row - 1, check_col).Value
            formatted_total_above = format_excel_style(total_above)
            logger.info(f"Total above at {cert_sheet.Cells(check_row - 1, check_col).Address}: {formatted_total_above}")

            # Get the total from the column before and one row above the check figure
            total_before = cert_sheet.Cells(check_row - 1, check_col - 1).Value
            formatted_total_before = format_excel_style(total_before)
            logger.info(f"Total before at {cert_sheet.Cells(check_row, check_col - 1).Address}: {formatted_total_before}")

            # Check for None values and convert to 0 if necessary
            check_figure = 0 if check_figure is None else check_figure
            total_above = 0 if total_above is None else total_above
            total_before = 0 if total_before is None else total_before

            # Sum these values
            try:
                cert_sum = check_figure + total_above + total_before
                formatted_cert_sum = format_excel_style(cert_sum)
                logger.info(f"Certification sum: {formatted_cert_sum}")
            except Exception as e:
                logger.error(f"Error calculating cert_sum: {str(e)}")
                logger.error(f"check_figure: {formatted_check_figure}, type: {type(check_figure)}")
                logger.error(f"total_above: {formatted_total_above}, type: {type(total_above)}")
                logger.error(f"total_before: {formatted_total_before}, type: {type(total_before)}")
                raise

            # Get the total from the Obligation Analysis sheet
            total_cell = obl_sheet.Cells(header_row + 7, last_column + 20)  # count_no_invoice_col
            obl_sum = total_cell.Value
            formatted_obl_sum = format_excel_style(obl_sum)
            logger.info(f"Obligation Analysis sum at {total_cell.Address}: {formatted_obl_sum}")

            if obl_sum is None:
                logger.warning("Obligation Analysis sum is None, setting to 0")
                obl_sum = 0

            logger.info(f"Comparing totals - Certification: {formatted_cert_sum}, Obligation Analysis: {formatted_obl_sum}")

            if abs(cert_sum - obl_sum) < 0.01:  # Using a small threshold for float comparison
                logger.info("Totals match")
                # Add 'h' tickmark to Certification sheet
                cert_tickmark_cell = cert_sheet.Cells(check_row, check_col + 1)
                cert_tickmark_cell.Font.Name = "Wingdings"
                cert_tickmark_cell.Font.Color = 0  # Black
                cert_tickmark_cell.Font.Size = 10
                cert_tickmark_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                cert_tickmark_cell.Value = "h"
                logger.info(f"Added 'h' tickmark to Certification sheet at {cert_tickmark_cell.Address}")

                # Add 'h' tickmark to Obligation Analysis sheet
                obl_tickmark_cell = obl_sheet.Cells(total_cell.Row, total_cell.Column + 1)
                obl_tickmark_cell.Font.Name = "Wingdings"
                obl_tickmark_cell.Font.Color = 0  # Black
                obl_tickmark_cell.Font.Size = 10
                obl_tickmark_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                obl_tickmark_cell.Value = "m"
                logger.info(f"Added 'h' tickmark to Obligation Analysis sheet at {obl_tickmark_cell.Address}")
            else:
                logger.warning("Totals do not match")
                # Add 'X' to Certification sheet
                cert_x_cell = cert_sheet.Cells(check_row, check_col + 1)
                cert_x_cell.Font.Name = "Calibri"
                cert_x_cell.Font.Color = 0  # Black
                cert_x_cell.Font.Size = 10
                cert_x_cell.Font.Bold = True
                cert_x_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                cert_x_cell.Value = "X"
                logger.info(f"Added 'X' to Certification sheet at {cert_x_cell.Address}")

                # Add 'X' to Obligation Analysis sheet
                obl_x_cell = obl_sheet.Cells(total_cell.Row, total_cell.Column + 1)
                obl_x_cell.Font.Name = "Calibri"
                obl_x_cell.Font.Color = 0  # Black
                obl_x_cell.Font.Size = 10
                obl_x_cell.Font.Bold = True
                obl_x_cell.HorizontalAlignment = win32com.client.constants.xlLeft
                obl_x_cell.Value = "X"
                logger.info(f"Added 'X' to Obligation Analysis sheet at {obl_x_cell.Address}")

    except Exception as e:
        logger.error(f"Error in comparing {table_name} tables: {str(e)}", exc_info=True)
        raise