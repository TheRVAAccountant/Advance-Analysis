"""
Complete data processing module for advance analysis.

This module provides the full implementation of data processing functionality
including CY processing, PY processing, merging, and StatusValidations.
"""

import pandas as pd
from datetime import datetime
from typing import Optional, Tuple
import os

from ..utils.logging_config import get_logger
from ..modules.data_loader import load_excel_file, find_header_row_in_dataframe
from .advance_analysis_processing import process_advance_analysis
from .comparative_analysis_processing import process_comparative_analysis
from .do_advance_analysis_processing import process_do_advance_analysis
from .data_processing_simple import get_comparative_period

logger = get_logger(__name__)


def find_comparative_file(directory: str, component: str, comparative_period: str) -> Optional[str]:
    """
    Find the comparative advance analysis file in the specified directory.
    
    Args:
        directory: Directory to search in
        component: Component code (e.g., "WMD")
        comparative_period: Comparative period (e.g., "FY24 Q3")
        
    Returns:
        Path to comparative file if found, None otherwise
    """
    import glob
    
    # Build search pattern
    # Look for files containing component, comparative period, and "Advance Analysis"
    patterns = [
        f"{component}*{comparative_period}*Advance Analysis*.xlsx",
        f"{component}*{comparative_period.replace(' ', '')}*Advance Analysis*.xlsx",
        f"{component}*{comparative_period.replace('FY', '')}*Advance Analysis*.xlsx"
    ]
    
    logger.info(f"Searching for comparative file in: {directory}")
    
    for pattern in patterns:
        search_path = os.path.join(directory, pattern)
        logger.debug(f"Trying pattern: {pattern}")
        files = glob.glob(search_path)
        
        if files:
            # Return the first match
            comparative_file = files[0]
            logger.info(f"Found comparative file: {os.path.basename(comparative_file)}")
            return comparative_file
    
    # If no file found with patterns, try a more general search
    all_files = glob.glob(os.path.join(directory, "*.xlsx"))
    for file in all_files:
        basename = os.path.basename(file)
        # Check if file contains component, period components, and "Advance Analysis"
        if (component in basename and 
            comparative_period.replace(" ", "") in basename.replace(" ", "") and 
            "Advance Analysis" in basename and
            "DO" not in basename):  # Exclude files with "DO" to avoid processed files
            logger.info(f"Found comparative file (general search): {basename}")
            return file
    
    return None


def process_complete_advance_analysis(
    advance_file_path: str,
    current_dhstier_path: str, 
    prior_dhstier_path: str,
    component: str,
    cy_fy_qtr: str,
    output_folder: str
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Process advance analysis data completely including CY, PY, and merged validations.
    
    Args:
        advance_file_path: Path to the advance analysis Excel file
        current_dhstier_path: Path to current year DHSTIER file
        prior_dhstier_path: Path to prior year DHSTIER file
        component: DHS component code
        cy_fy_qtr: Current fiscal year and quarter
        output_folder: Path to output folder for saving results
        
    Returns:
        Tuple of (cy_df, py_df, merged_df) DataFrames
    """
    logger.info(f"Starting complete advance analysis processing for {component} {cy_fy_qtr}")
    
    # Extract fiscal year and calculate dates
    fiscal_year = int(cy_fy_qtr[2:4])
    current_reporting_date = get_reporting_date(cy_fy_qtr)
    fiscal_year_start_date = datetime(2000 + fiscal_year - 1, 10, 1)
    fiscal_year_end_date = datetime(2000 + fiscal_year, 9, 30)
    
    logger.info(f"Fiscal Year: 20{fiscal_year}")
    logger.info(f"Current Reporting Date: {current_reporting_date.strftime('%m/%d/%Y')}")
    logger.info(f"Fiscal Year Start: {fiscal_year_start_date.strftime('%m/%d/%Y')}")
    logger.info(f"Fiscal Year End: {fiscal_year_end_date.strftime('%m/%d/%Y')}")
    
    # Step 1: Load and process current year data
    logger.info("=" * 80)
    logger.info("STEP 1: Processing Current Year (CY) Data")
    logger.info("=" * 80)
    
    cy_df = load_excel_file(advance_file_path, sheet_name="4-Advance Analysis")
    cy_df = process_advance_analysis(
        df=cy_df,
        component=component,
        current_reporting_date=current_reporting_date,
        fiscal_year_start_date=fiscal_year_start_date
    )
    
    logger.info(f"CY processing complete. Shape: {cy_df.shape}")
    logger.info(f"CY columns: {list(cy_df.columns)}")
    
    # Save CY data
    cy_output_path = os.path.join(output_folder, "CY_Advance_Analysis.xlsx")
    cy_df.to_excel(cy_output_path, index=False)
    logger.info(f"CY data saved to: {cy_output_path}")
    
    # Step 2: Load and process prior year data
    logger.info("=" * 80)
    logger.info("STEP 2: Processing Prior Year (PY) Data")
    logger.info("=" * 80)
    
    # Calculate comparative period
    quarter = cy_fy_qtr[-2:]
    comparative_period = get_comparative_period(fiscal_year, quarter)
    logger.info(f"Comparative Period: {comparative_period}")
    
    # Find comparative file in the same directory as the advance file
    advance_dir = os.path.dirname(advance_file_path)
    
    # Search for comparative file
    comparative_file_path = find_comparative_file(advance_dir, component, comparative_period)
    
    if not comparative_file_path:
        error_msg = f"Comparative file not found for {component} {comparative_period} in directory: {advance_dir}"
        logger.error(error_msg)
        logger.error("Please ensure the comparative period advance analysis file is in the same directory as the current period file.")
        logger.error(f"Expected file pattern: {component}*{comparative_period}*Advance Analysis*.xlsx")
        raise FileNotFoundError(error_msg)
    
    logger.info(f"Found comparative file: {comparative_file_path}")
    
    # Load PY data from comparative file
    py_df = load_excel_file(comparative_file_path, sheet_name="4-Advance Analysis")
    py_df = process_comparative_analysis(
        df=py_df,
        component=component
    )
    
    logger.info(f"PY processing complete. Shape: {py_df.shape}")
    logger.info(f"PY columns: {list(py_df.columns)}")
    
    # Save PY data
    py_output_path = os.path.join(output_folder, "PY_Advance_Analysis.xlsx")
    py_df.to_excel(py_output_path, index=False)
    logger.info(f"PY data saved to: {py_output_path}")
    
    # Step 3: Merge and apply DO advance analysis validations
    logger.info("=" * 80)
    logger.info("STEP 3: Merging CY and PY Data with StatusValidations")
    logger.info("=" * 80)
    
    merged_df = process_do_advance_analysis(
        cy_df=cy_df,
        py_df=py_df,
        component=component,
        fiscal_year_start_date=fiscal_year_start_date,
        fiscal_year_end_date=fiscal_year_end_date
    )
    
    logger.info(f"Merged processing complete. Shape: {merged_df.shape}")
    logger.info(f"Merged columns: {list(merged_df.columns)}")
    
    # Log sample of merged data
    logger.info("=" * 80)
    logger.info("SAMPLE OF MERGED DATA (First 5 rows):")
    logger.info("=" * 80)
    
    # Select key columns for logging
    key_columns = [
        'DO Concatenate', 'Status', 'Advance/Prepayment',
        'Advances Requiring Explanations?', 'Null or Blank Columns',
        'Status Changed?', 'Valid Status 1', 'Valid Status 2',
        'DO Status 1 Validation', 'DO Status 2 Validations', 'DO Comment'
    ]
    
    available_columns = [col for col in key_columns if col in merged_df.columns]
    sample_df = merged_df[available_columns].head()
    logger.info(f"\n{sample_df.to_string()}")
    
    # Log validation statistics
    logger.info("=" * 80)
    logger.info("VALIDATION STATISTICS:")
    logger.info("=" * 80)
    
    if 'Advances Requiring Explanations?' in merged_df.columns:
        exp_stats = merged_df['Advances Requiring Explanations?'].value_counts()
        logger.info(f"Advances Requiring Explanations:\n{exp_stats}")
    
    if 'Status Changed?' in merged_df.columns:
        status_change_stats = merged_df['Status Changed?'].value_counts()
        logger.info(f"Status Changed:\n{status_change_stats}")
    
    if 'Valid Status 1' in merged_df.columns:
        valid1_stats = merged_df['Valid Status 1'].value_counts()
        logger.info(f"Valid Status 1:\n{valid1_stats}")
    
    if 'Valid Status 2' in merged_df.columns:
        valid2_stats = merged_df['Valid Status 2'].value_counts()
        logger.info(f"Valid Status 2:\n{valid2_stats}")
    
    # Save merged data with all validations
    merged_output_path = os.path.join(output_folder, "DO_Tab_4_Review_Data.xlsx")
    
    # Format date columns before saving
    date_columns = [
        'Date of Advance', 'Last Activity Date', 'Anticipated Liquidation Date',
        'Period of Performance End Date', 'Date of Advance_comp', 
        'Last Activity Date_comp', 'Anticipated Liquidation Date_comp',
        'Period of Performance End Date_comp'
    ]
    
    for col in date_columns:
        if col in merged_df.columns:
            merged_df[col] = pd.to_datetime(merged_df[col], errors='coerce')
    
    # Save with date formatting - ensure proper file closure for COM access
    try:
        # Use openpyxl directly for better control over file saving
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import NamedStyle
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'DO Tab 4 Review'
        
        # Write dataframe to worksheet
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws.append(r)
        
        # Apply date formatting
        date_style = NamedStyle(name='date_style')
        date_style.number_format = '*m/dd/yyyy'
        
        # Find and format date columns
        headers = [cell.value for cell in ws[1]]
        for col_idx, col_name in enumerate(headers, 1):
            if col_name in date_columns:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, datetime):
                        cell.style = date_style
        
        # Save the workbook
        wb.save(merged_output_path)
        wb.close()
        
        logger.info(f"Merged data with validations saved to: {merged_output_path}")
        
        # Ensure file is fully written to disk
        import time
        time.sleep(0.2)  # Small delay for file system
        
        # Verify file exists and is accessible
        if not os.path.exists(merged_output_path):
            raise FileNotFoundError(f"Output file was not created: {merged_output_path}")
            
        file_size = os.path.getsize(merged_output_path)
        logger.info(f"Output file size: {file_size} bytes")
        
        # Verify file is not locked by trying to open it
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                with open(merged_output_path, 'rb') as f:
                    # Read first few bytes to ensure it's accessible
                    f.read(100)
                logger.info("Verified output file is accessible and not locked")
                break
            except Exception as e:
                if attempt < max_attempts - 1:
                    logger.warning(f"Attempt {attempt + 1}/{max_attempts} - File may be locked: {e}. Retrying...")
                    time.sleep(0.5)
                else:
                    logger.error(f"Output file appears to be locked after {max_attempts} attempts: {e}")
                    raise
            
    except Exception as e:
        logger.error(f"Error saving merged data: {e}")
        raise
    
    return cy_df, py_df, merged_df


def get_reporting_date(cy_fy_qtr: str) -> datetime:
    """
    Determine the reporting date based on fiscal year and quarter.
    
    Args:
        cy_fy_qtr (str): Current fiscal year and quarter (e.g., "FY24 Q2").
    
    Returns:
        datetime: The reporting date.
        
    Raises:
        ValueError: If the quarter format is invalid.
    """
    fiscal_year = int(cy_fy_qtr[2:4])
    quarter = cy_fy_qtr[-2:]

    if quarter == 'Q1':
        return datetime(2000 + fiscal_year - 1, 12, 31)
    elif quarter == 'Q2':
        return datetime(2000 + fiscal_year, 3, 31)
    elif quarter == 'Q3':
        return datetime(2000 + fiscal_year, 6, 30)
    elif quarter == 'Q4':
        return datetime(2000 + fiscal_year, 9, 30)
    else:
        raise ValueError(f"Invalid quarter format: {quarter}")