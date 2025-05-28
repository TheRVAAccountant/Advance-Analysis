import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_test_excel_files():
    # Define sheet names
    sheet_names = [
        "1-Instructions",
        "2-Certification",
        "3-PY Q4 Ending Balance",
        "4-Advance Analysis",
        "5-FY18 Scorecard Metrics",
        "6-ADVANCE TO TIER Recon SUMMARY",
        "Version Control",
        "DO USE Only"
    ]
    
    # Create main headers for 4-Advance Analysis sheet
    advance_analysis_headers = [
        "TAS", "SGL", "DHS Doc No", "Indicate if advance is to WCF (Y/N)",
        "Advance/Prepayment", "Date of Advance", "Age of Advance (days)",
        "Last Activity Date", "Anticipated Liquidation Date",
        "Period of Performance End Date", "Status", "Advance/Prepayment_1",
        "Comments", "Vendor", "Trading Partner ID",
        "Advance Type (e.g. Travel, Vendor Prepayment)"
    ]
    
    # Create sample data for main file (FY25 Q2)
    main_data = [
        ["70-0715", "1410", "HSSCCG23F00001", "N", 250000.00, 
         datetime(2024, 10, 15), 104, datetime(2024, 12, 28),
         datetime(2025, 3, 31), datetime(2025, 9, 30), "Open",
         250000.00, "Q2 advance payment", "ABC Contractor", "1234567890",
         "Vendor Prepayment"],
        ["70-0715", "1410", "HSSCCG23F00002", "Y", 150000.00,
         datetime(2024, 11, 1), 87, datetime(2024, 12, 15),
         datetime(2025, 2, 28), datetime(2025, 6, 30), "Open",
         150000.00, "Working capital fund advance", "XYZ Services", "0987654321",
         "Travel"],
        ["70-0715", "1410", "HSSCCG23F00003", "N", 75000.00,
         datetime(2024, 9, 30), 119, datetime(2024, 12, 1),
         datetime(2025, 1, 31), datetime(2025, 3, 31), "Partially Liquidated",
         50000.00, "Partial liquidation in progress", "DEF Corp", "5555555555",
         "Vendor Prepayment"],
        ["70-0715", "1410", "HSSCCG23F00004", "N", 100000.00,
         datetime(2024, 8, 15), 165, datetime(2024, 11, 30),
         datetime(2024, 12, 31), datetime(2025, 2, 28), "Open",
         100000.00, "Extended due to project delays", "GHI Solutions", "9999999999",
         "Services"]
    ]
    
    # Create sample data for comparative file (FY24 Q3)
    comp_data = [
        ["70-0715", "1410", "HSSCCG23F00001", "N", 250000.00,
         datetime(2024, 10, 15), 15, datetime(2024, 10, 30),
         datetime(2025, 3, 31), datetime(2025, 9, 30), "Open",
         250000.00, "Initial advance payment", "ABC Contractor", "1234567890",
         "Vendor Prepayment"],
        ["70-0715", "1410", "HSSCCG23F00003", "N", 75000.00,
         datetime(2024, 9, 30), 30, datetime(2024, 10, 15),
         datetime(2025, 1, 31), datetime(2025, 3, 31), "Open",
         75000.00, "New advance", "DEF Corp", "5555555555",
         "Vendor Prepayment"],
        ["70-0715", "1410", "HSSCCG22F00099", "N", 50000.00,
         datetime(2023, 12, 15), 320, datetime(2024, 9, 30),
         datetime(2024, 6, 30), datetime(2024, 9, 30), "Liquidated",
         0.00, "Fully liquidated in Q3", "Old Vendor", "1111111111",
         "Travel"]
    ]
    
    # PY Q4 Ending Balance headers
    py_balance_headers = [
        "TAS", "DHS Doc No", "Advance/Prepayment", "Advance/Prepayment",
        "Status", "Date of Advance", "Date of Advance - Age", "Last Activity Date"
    ]
    
    # PY Q4 sample data
    py_balance_data = [
        ["70-0715", "HSSCCG23F00001", 250000.00, 250000.00, "Open",
         datetime(2024, 10, 15), 104, datetime(2024, 12, 28)],
        ["70-0715", "HSSCCG23F00003", 75000.00, 50000.00, "Partially Liquidated",
         datetime(2024, 9, 30), 119, datetime(2024, 12, 1)]
    ]
    
    # Create FY25 Q2 file
    create_excel_file(
        "/Users/jeroncrooks/CascadeProjects/Advance-Analysis/test_data/WMD FY25 Q2 Advance Analysis Test.xlsx",
        sheet_names,
        advance_analysis_headers,
        main_data,
        py_balance_headers,
        py_balance_data,
        "FY25 Q2"
    )
    
    # Create FY24 Q3 file
    create_excel_file(
        "/Users/jeroncrooks/CascadeProjects/Advance-Analysis/test_data/WMD FY24 Q3 Advance Analysis Test.xlsx",
        sheet_names,
        advance_analysis_headers,
        comp_data,
        py_balance_headers,
        py_balance_data[:1],  # Less data for comparative period
        "FY24 Q3"
    )

def create_excel_file(filename, sheet_names, adv_headers, adv_data, py_headers, py_data, period):
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)
        
        if sheet_name == "1-Instructions":
            ws['A1'] = f"Advance Analysis Instructions for {period}"
            ws['A3'] = "This is a test file for the Advance Analysis system"
            
        elif sheet_name == "2-Certification":
            ws['A1'] = f"Certification for {period}"
            ws['B10'] = "Advances"
            ws['B11'] = "$-823,769.84"
            
        elif sheet_name == "3-PY Q4 Ending Balance":
            # Add headers at row 6
            for col, header in enumerate(py_headers, 1):
                ws.cell(row=6, column=col, value=header)
            
            # Add data
            for row_idx, row_data in enumerate(py_data, 7):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
        elif sheet_name == "4-Advance Analysis":
            # Add some blank rows before headers
            for i in range(1, 10):
                ws.cell(row=i, column=1, value="")
            
            # Add headers at row 10
            for col, header in enumerate(adv_headers, 1):
                ws.cell(row=10, column=col, value=header)
            
            # Add data starting at row 11
            for row_idx, row_data in enumerate(adv_data, 11):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
        elif sheet_name == "5-FY18 Scorecard Metrics":
            ws['A1'] = "Scorecard Metrics"
            ws['A3'] = "Test metrics data"
            
        elif sheet_name == "6-ADVANCE TO TIER Recon SUMMARY":
            ws['A1'] = "Reconciliation Summary"
            ws['A3'] = "Test reconciliation data"
            
        elif sheet_name == "Version Control":
            ws['A1'] = "Version Control"
            ws['A3'] = f"Test file created for {period}"
            ws['A4'] = f"Created on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
        elif sheet_name == "DO USE Only":
            ws['A1'] = "DO USE Only"
            ws['A3'] = "Internal use section"
    
    # Save the workbook
    wb.save(filename)
    print(f"Created: {filename}")

if __name__ == "__main__":
    create_test_excel_files()
    print("Test files created successfully!")